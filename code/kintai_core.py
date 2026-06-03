"""勤務表画像・PDF解析（NewtonX）のコア処理。"""
from __future__ import annotations

from collections.abc import Callable

from newtonx_adk import NewtonXClient, ConfigManager, FileUploadError
from openpyxl import load_workbook
import json
import threading
import math
import re
import tempfile
import time
import types
import unicodedata
from datetime import datetime as DTDateTime, time as DTTime, timedelta as DTTimeDelta
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

from PIL import Image, ImageOps, UnidentifiedImageError

# バイナリ換算の 1MiB。「未満」なのでこれ未満のサイズになるよう調整する
MIB = 1024 * 1024
# PDF をラスタ化する際の解像度（dpi）
PDF_RENDER_DPI = 150

IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".tif", ".tiff"}
PDF_SUFFIX = ".pdf"
EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
TARGET_EXCEL_SHEET_NAME = "タイムシート兼作業報告書_お客様先用"
TRANSPORT_EXPENSE_EXCEL_SHEET_NAME = "交通宿泊費清算書お客様先用"
TRANSPORT_EXPENSE_LABEL_COL = 6  # F列: 「合計」ラベル
TRANSPORT_EXPENSE_VALUE_COL = 7  # G列: 交通費合計の値
TRANSPORT_EXPENSE_LABEL_TEXT = "合計"
SUMMARY_TRANSPORT_EXPENSE_COL = "交通費合計（読取）"

# アップロード: 初回試行後、最大 UPLOAD_MAX_RETRIES 回まで再試行（合計で最大 1 + UPLOAD_MAX_RETRIES 回）
UPLOAD_MAX_RETRIES = 3
UPLOAD_RETRY_DELAY_SEC = 0.35
# 解析1件ごとの delete_chat を行うか（False: チャットは NewtonX 上に残す）
DELETE_CHAT_AFTER_ANALYSIS = False

DEFAULT_PARALLEL_ANALYSIS_CHATS = 1
PARALLEL_WORKERS_MAX = 32

# グリッド・集計表の先頭列（〇: リトライ上限内でアップロード成功 / ✖: それ以外）
SUMMARY_UPLOAD_COL = "アップロード"
SUMMARY_TARGET_SHEET_COL = "対象シート有無"
TARGET_FILE_NAME_COL = "対象ファイル名（左クリックで表示）"
LEGACY_TARGET_FILE_NAME_COL = "対象ファイル名"
LEGACY_FILE_NAME_COL = "画像ファイル名"
SUMMARY_COMPANY_COL = "会社名（読取）"
LEGACY_COMPANY_READ_LONG_COL = "会社名（読み取）"
LEGACY_COMPANY_COL = "会社名1"
LEGACY_COMPANY_NAME_COL = "会社名"
SUMMARY_YEAR_COL = "年（読取）"
SUMMARY_MONTH_COL = "月（読取）"
LEGACY_YEAR_COL = "年"
LEGACY_MONTH_COL = "月"
SUMMARY_PERSON_COL = "氏名（読取）"
LEGACY_PERSON_COL = "氏名"
SUMMARY_EMPLOYEE_NO_COL = "社員番号（ファイル名より）"
LEGACY_EMPLOYEE_NO_COL = "社員番号"
SUMMARY_ROW_NO_COL = "No"
SUMMARY_BILLING_UPDATE_HOURS_COL = "更新用合計勤務時間（10進）"
LEGACY_BILLING_UPDATE_HOURS_COL = "更新用合計時間（10進数）"
SUMMARY_BILLING_UPDATE_TRANSPORT_COL = "更新用交通費合計"
SUMMARY_FINAL_JUDGMENT_COL = "最終判断"
LEGACY_USER_JUDGMENT_COL = "ユーザ判断"
SUMMARY_BILLING_UPDATE_RESULT_COL = "請求用ファイル更新"
LEGACY_BILLING_UPDATE_RESULT_COL = "請求量ファイル更新結果"
BILLING_UPDATE_NO_TARGET_RECORD = "（対象レコードなし）"
BILLING_ENGINEER_TS_SHEET_NAME = "エンジニアTS一覧"
BILLING_TS_COL_EMPLOYEE_NO = 1  # A列
BILLING_TS_COL_CLIENT_NAME = 4  # D列: 取引先名
BILLING_TS_COL_PERSON_NAME = 5  # E列: 氏名
BILLING_TS_COL_NORMAL_HOURS = 8  # H列: 通常請求時間
BILLING_TS_COL_TRANSPORT_AMOUNT = 15  # O列: 旅費交通費請求金額

#TARGET_ASSISTANT_NAME = "GPT-5.2(高性能)"
TARGET_ASSISTANT_NAME = "GPT-5.4-mini(高速)"
#TARGET_ASSISTANT_NAME = "Gemini 3.1 Pro(高性能)"

def _process_sse_response_no_print(self, response) -> str:
    full_response = ""
    for line in response.iter_lines(decode_unicode=True):
        if line and line.startswith("data: "):
            data = line[6:]
            if data == "[DONE]":
                break
            try:
                parsed_data = json.loads(data)
                if parsed_data.get("type") == "content":
                    full_response += parsed_data.get("content", "")
            except json.JSONDecodeError:
                full_response += data
    return full_response


def create_client() -> NewtonXClient:
    """NewtonX クライアントを生成し、SSE の標準出力を抑制するパッチを当てる。"""
    config_manager = ConfigManager()
    _cfg = config_manager.get_config()
    _cfg.timeout = max(getattr(_cfg, "timeout", 30) or 30, 300)
    client = NewtonXClient(config_manager)
    client._process_sse_response = types.MethodType(
        _process_sse_response_no_print, client
    )
    return client


def _compress_image_under_1mib(src: Path) -> Path:
    """元ファイルは変更しない。1MiB未満のJPEGを tempfile に書き、そのパスを返す。"""
    img = Image.open(src)
    img.load()
    img = ImageOps.exif_transpose(img)

    if getattr(img, "is_animated", False):
        img.seek(0)

    if img.mode in ("RGBA", "LA"):
        bg = Image.new("RGB", img.size, (255, 255, 255))
        bg.paste(img, mask=img.split()[-1])
        img = bg
    elif img.mode == "P":
        rgba = img.convert("RGBA")
        bg = Image.new("RGB", rgba.size, (255, 255, 255))
        bg.paste(rgba, mask=rgba.split()[-1])
        img = bg
    elif img.mode != "RGB":
        img = img.convert("RGB")

    tmp = tempfile.NamedTemporaryFile(prefix="nx_upload_", suffix=".jpg", delete=False)
    tmp.close()
    out_path = Path(tmp.name)

    current = img
    for _ in range(20):
        for quality in range(92, 9, -4):
            current.save(out_path, format="JPEG", quality=quality, optimize=True)
            if out_path.stat().st_size < MIB:
                return out_path
        w, h = current.size
        if w <= 320 and h <= 320:
            break
        current = current.resize(
            (max(1, int(w * 0.88)), max(1, int(h * 0.88))),
            Image.Resampling.LANCZOS,
        )

    raise RuntimeError(f"1MiB未満に収められませんでした: {src}")


def _otsu_threshold(gray: Image.Image) -> int:
    """グレースケール画像のヒストグラムから大津の二値化閾値を求める。"""
    hist = gray.histogram()
    total = sum(hist)
    if total == 0:
        return 128
    sum_total = sum(i * h for i, h in enumerate(hist))
    sum_b = 0
    w_b = 0
    max_var = 0.0
    threshold = 128
    for t in range(256):
        w_b += hist[t]
        if w_b == 0:
            continue
        w_f = total - w_b
        if w_f == 0:
            break
        sum_b += t * hist[t]
        m_b = sum_b / w_b
        m_f = (sum_total - sum_b) / w_f
        var_between = w_b * w_f * (m_b - m_f) ** 2
        if var_between > max_var:
            max_var = var_between
            threshold = t
    return threshold


def _binarize_image_monochrome(img: Image.Image) -> Image.Image:
    """RGB/グレースケール画像をモノクロ2値（白黒のみ）にする。"""
    gray = img.convert("L")
    threshold = _otsu_threshold(gray)
    return gray.point(lambda p: 255 if p > threshold else 0, mode="L")


def _strip_pdf_structure_tree(doc) -> None:
    """破損したPDFタグ（Structure Tree）をメモリ上で無効化する。

    元ファイルは変更しない。get_pixmap が遅くなる／MuPDFエラーが出るPDFへの対処。
    """
    try:
        pdf_catalog = doc.pdf_catalog
        cat = pdf_catalog() if callable(pdf_catalog) else pdf_catalog
        stree = doc.xref_get_key(cat, "StructTreeRoot")
        if stree[1] != "null":
            doc.xref_set_key(cat, "StructTreeRoot", "null")
    except Exception:
        pass


def _resolve_upload_path(original: Path) -> tuple[str, Path | None]:
    """アップロード用パスと、削除すべき一時ファイル（あれば）を返す。元ファイルは書き換えない。"""
    if original.stat().st_size < MIB:
        return str(original), None
    tmp = _compress_image_under_1mib(original)
    return str(tmp), tmp


def _pdf_to_png_for_upload(pdf_path: Path) -> tuple[str, list[Path]]:
    """PDFをPNGに変換し、1MiB未満のアップロード用パスを返す。

    各ページはモノクロ2値化（大津の方法）後、複数ページは縦に結合する。
    1MiB以上のPNGは既存の画像圧縮でJPEG化する。
    戻り値: (upload_path, 削除すべき一時ファイル一覧)
    """
    try:
        import fitz
    except ImportError as e:
        raise RuntimeError(
            "PDFを画像に変換するには pymupdf が必要です（pip install pymupdf）"
        ) from e

    tools = fitz.TOOLS
    tools.mupdf_display_errors(False)
    tools.mupdf_display_warnings(False)
    try:
        if hasattr(tools, "mupdf_warnings"):
            tools.mupdf_warnings(reset=True)

        temps: list[Path] = []
        doc = fitz.open(pdf_path)
        try:
            _strip_pdf_structure_tree(doc)
            page_count = doc.page_count
            if page_count < 1:
                raise ValueError(f"ページがありません: {pdf_path.name}")

            zoom = PDF_RENDER_DPI / 72.0
            mat = fitz.Matrix(zoom, zoom)
            page_images: list[Image.Image] = []
            for page_idx in range(page_count):
                pix = doc[page_idx].get_pixmap(matrix=mat, alpha=False)
                page_img = Image.frombytes(
                    "RGB", (pix.width, pix.height), pix.samples
                )
                page_images.append(_binarize_image_monochrome(page_img))

            if len(page_images) == 1:
                combined = page_images[0]
            else:
                width = max(im.width for im in page_images)
                height = sum(im.height for im in page_images)
                combined = Image.new("L", (width, height), 255)
                y_off = 0
                for im in page_images:
                    combined.paste(im, (0, y_off))
                    y_off += im.height

            png_tmp = tempfile.NamedTemporaryFile(
                prefix="nx_pdf_", suffix=".png", delete=False
            )
            png_tmp.close()
            png_path = Path(png_tmp.name)
            temps.append(png_path)
            combined.save(png_path, format="PNG", optimize=True)

            upload_src, compressed_tmp = _resolve_upload_path(png_path)
            if compressed_tmp is not None:
                temps.append(compressed_tmp)
            return upload_src, temps
        finally:
            doc.close()
    finally:
        tools.mupdf_display_errors(True)
        tools.mupdf_display_warnings(True)


def _filename_has_transport_expense_marker(file_name: str) -> bool:
    """ファイル名に「交通費」が含まれるか（全角半角正規化後）。"""
    return "交通費" in unicodedata.normalize("NFKC", file_name or "")


def _transport_expense_prompt_block(display_file_name: str) -> str:
    """ファイル名に交通費を含む場合のみ、AI への交通費合計抽出指示を返す。"""
    if not _filename_has_transport_expense_marker(display_file_name):
        return ""
    return """
４）交通費合計の抽出（ファイル名に「交通費」を含むため）
交通費・経費精算の書類から、交通費等の合計と思われる金額を抽出し「交通費合計」として出力してください。
税抜き・税込みの両方の金額がある場合は、税込みの金額を採用してください。
金額は表記のまま（円・カンマ等を含む）で構いません。読み取れない場合は「（なし）」。
"""


# NewtonX 解析応答 JSON のキー名（仕様固定）
ANALYSIS_JSON_FIELD_TARGET_FILE = "対象ファイル"
ANALYSIS_JSON_FIELD_COMPANY = "会社名"
ANALYSIS_JSON_FIELD_PERSON = "氏名"
ANALYSIS_JSON_FIELD_YEAR = "年度"
ANALYSIS_JSON_FIELD_MONTH = "月"
ANALYSIS_JSON_FIELD_TOTAL_HOURS = "合計勤務時間"
ANALYSIS_JSON_FIELD_SEAL = "押印有無"
ANALYSIS_JSON_FIELD_TRANSPORT = "交通費合計"

ANALYSIS_JSON_CANONICAL_KEYS: frozenset[str] = frozenset(
    {
        ANALYSIS_JSON_FIELD_TARGET_FILE,
        ANALYSIS_JSON_FIELD_COMPANY,
        ANALYSIS_JSON_FIELD_PERSON,
        ANALYSIS_JSON_FIELD_YEAR,
        ANALYSIS_JSON_FIELD_MONTH,
        ANALYSIS_JSON_FIELD_TOTAL_HOURS,
        ANALYSIS_JSON_FIELD_SEAL,
        ANALYSIS_JSON_FIELD_TRANSPORT,
    }
)

_ANALYSIS_JSON_KEY_ALIASES: dict[str, str] = {
    "対象ファイル": ANALYSIS_JSON_FIELD_TARGET_FILE,
    "対象ファイル名": ANALYSIS_JSON_FIELD_TARGET_FILE,
    TARGET_FILE_NAME_COL: ANALYSIS_JSON_FIELD_TARGET_FILE,
    LEGACY_TARGET_FILE_NAME_COL: ANALYSIS_JSON_FIELD_TARGET_FILE,
    LEGACY_FILE_NAME_COL: ANALYSIS_JSON_FIELD_TARGET_FILE,
    "ファイル名": ANALYSIS_JSON_FIELD_TARGET_FILE,
    "file_name": ANALYSIS_JSON_FIELD_TARGET_FILE,
    "会社名": ANALYSIS_JSON_FIELD_COMPANY,
    "company": ANALYSIS_JSON_FIELD_COMPANY,
    "氏名": ANALYSIS_JSON_FIELD_PERSON,
    "person": ANALYSIS_JSON_FIELD_PERSON,
    "年度": ANALYSIS_JSON_FIELD_YEAR,
    "年": ANALYSIS_JSON_FIELD_YEAR,
    "year": ANALYSIS_JSON_FIELD_YEAR,
    "月": ANALYSIS_JSON_FIELD_MONTH,
    "month": ANALYSIS_JSON_FIELD_MONTH,
    "合計勤務時間": ANALYSIS_JSON_FIELD_TOTAL_HOURS,
    "total_hours": ANALYSIS_JSON_FIELD_TOTAL_HOURS,
    "押印有無": ANALYSIS_JSON_FIELD_SEAL,
    "押印": ANALYSIS_JSON_FIELD_SEAL,
    "seal": ANALYSIS_JSON_FIELD_SEAL,
    "交通費合計": ANALYSIS_JSON_FIELD_TRANSPORT,
    "交通費": ANALYSIS_JSON_FIELD_TRANSPORT,
    "transport": ANALYSIS_JSON_FIELD_TRANSPORT,
}


def _normalize_analysis_json_object(obj: object) -> dict[str, str] | None:
    """解析応答 JSON オブジェクトを正規化した辞書に変換する。"""
    if not isinstance(obj, dict):
        return None
    out: dict[str, str] = {k: "" for k in ANALYSIS_JSON_CANONICAL_KEYS}
    for raw_key, raw_val in obj.items():
        key = unicodedata.normalize("NFKC", str(raw_key or "").strip())
        if not key:
            continue
        canon = _ANALYSIS_JSON_KEY_ALIASES.get(key, key)
        if canon not in ANALYSIS_JSON_CANONICAL_KEYS:
            continue
        if raw_val is None:
            out[canon] = ""
        else:
            out[canon] = str(raw_val).strip()
    if not any(out.values()):
        return None
    return out


def _parse_analysis_json_response(text: str) -> dict[str, str] | None:
    """LLM 応答文字列から解析結果 JSON を抽出する。"""
    raw = (text or "").strip()
    if not raw:
        return None
    candidates: list[str] = [raw]
    fence = re.search(r"```(?:json)?\s*([\s\S]*?)```", raw, re.IGNORECASE)
    if fence:
        candidates.insert(0, fence.group(1).strip())
    brace = re.search(r"\{[\s\S]*\}", raw)
    if brace:
        candidates.append(brace.group(0))
    for candidate in candidates:
        if not candidate:
            continue
        try:
            obj = json.loads(candidate)
        except json.JSONDecodeError:
            continue
        normalized = _normalize_analysis_json_object(obj)
        if normalized:
            return normalized
    return None


def _print_analysis_json_to_terminal(
    file_name: str,
    analysis_json: dict[str, str] | None,
    raw_response: str,
) -> None:
    """生成AI応答の JSON をターミナル（stdout）に出力する。"""
    label = (file_name or "").strip() or "（不明）"
    if analysis_json:
        body = json.dumps(analysis_json, ensure_ascii=False, indent=2)
        print(f"--- 解析JSON: {label} ---\n{body}\n---", flush=True)
        return
    raw = (raw_response or "").strip()
    if raw:
        print(f"--- 解析JSON（パース失敗）: {label} ---\n{raw}\n---", flush=True)


def _analysis_json_to_kintai_tab(analysis_json: dict[str, str]) -> dict[str, str]:
    """解析 JSON を _enrich_with_match_scores 用の内部キー辞書に変換する。"""
    return {
        "company": (analysis_json.get(ANALYSIS_JSON_FIELD_COMPANY) or "").strip(),
        "person": (analysis_json.get(ANALYSIS_JSON_FIELD_PERSON) or "").strip(),
        "year": (analysis_json.get(ANALYSIS_JSON_FIELD_YEAR) or "").strip(),
        "month": (analysis_json.get(ANALYSIS_JSON_FIELD_MONTH) or "").strip(),
        "total": (analysis_json.get(ANALYSIS_JSON_FIELD_TOTAL_HOURS) or "").strip(),
        "seal": (analysis_json.get(ANALYSIS_JSON_FIELD_SEAL) or "").strip(),
        "transport": (analysis_json.get(ANALYSIS_JSON_FIELD_TRANSPORT) or "").strip(),
    }


def _build_json_response_instruction(display_file_name: str) -> str:
    """画像・PDF 共通: JSON 出力形式の指示ブロック。"""
    transport_note = ""
    if _filename_has_transport_expense_marker(display_file_name):
        transport_note = (
            f'  "{ANALYSIS_JSON_FIELD_TRANSPORT}": "税込み優先。読取不可は（なし）",\n'
        )
    else:
        transport_note = f'  "{ANALYSIS_JSON_FIELD_TRANSPORT}": "",\n'
    return f"""
７）出力形式
必ず次のキー名を持つ JSON オブジェクト1つのみを返してください。
説明文・Markdown 表・コードフェンス（```）は付けないでください。値はすべて文字列にしてください。

{{
  "{ANALYSIS_JSON_FIELD_TARGET_FILE}": "{display_file_name}",
  "{ANALYSIS_JSON_FIELD_COMPANY}": "",
  "{ANALYSIS_JSON_FIELD_PERSON}": "",
  "{ANALYSIS_JSON_FIELD_YEAR}": "",
  "{ANALYSIS_JSON_FIELD_MONTH}": "",
  "{ANALYSIS_JSON_FIELD_TOTAL_HOURS}": "",
  "{ANALYSIS_JSON_FIELD_SEAL}": "〇 または ✖",
{transport_note}}}
"""


def _apply_analysis_response_to_row(
    row: dict[str, str],
    analysis_json: dict[str, str] | None,
    raw_response: str,
    *,
    prefer_kintai_section: bool = False,
) -> None:
    """analyze_image_once / analyze_pdf_once の戻り値を行データに反映する。"""
    _print_analysis_json_to_terminal(
        row.get("file_name") or "",
        analysis_json,
        raw_response,
    )
    if analysis_json:
        row["analysis"] = json.dumps(analysis_json, ensure_ascii=False, indent=2)
    else:
        row["analysis"] = (
            raw_response if raw_response else "（解析結果を取得できませんでした）"
        )
    _enrich_with_match_scores(
        row,
        row["analysis"],
        prefer_kintai_section=prefer_kintai_section,
        analysis_json=analysis_json,
    )


def _build_check_message(display_file_name: str) -> str:
    return f"""
    	アップロードした画像ファイルから以下の情報を抽出してJSON形式でデータのみ返信してください。
    	
    	・対象ファイル
		・会社名
		・氏名
		・年度
		・月
		・合計勤務時間
		・押印有無
		・交通費合計

        【重要ルール】
        - 画像内に文字として明確に確認できない情報は、いかなる場合も推測・補完・類推して返さないこと。
        - 会社名、氏名は、検索キーワードに近いからという理由で補完してはならない。
        - 画像から明瞭に読み取れた文字列のみを返すこと。
        - 1文字でも判読が不確かな場合は「（データなし）」とする。
        - 画像内に存在しない、または確認できない会社名・氏名を返してはならない。

    	各項目の説明
    	1)対象ファイル
    		{display_file_name}を値として返す。
    	2)会社名
    		アップロードした画像ファイル内の文字をOCRし、画像内に明確に記載された会社名のみを返してください。
            会社検索キーワードはあくまで探索の補助であり、JSONの返却値ではありません。
            キーワードに類似していても、画像内に明確な記載が確認できない場合は「（データなし）」を返してください。
            会社検索キーワードを返すのではなく、画像から明確に抽出できた会社名を返してください。
            読み取れない、または確証がない場合は「（データなし）」を返してください。
            会社検索キーワードというものを以下に設定するがこれは返却値ではない。
    		{display_file_name}の先頭に[]でくくられた文字列がある場合、それ以降から次の'_’までの文字列を会社検索キーワードとする（最後の様などの継承は除く）。
    		{display_file_name}の先頭に[]でくくられた部分がない場合は、先頭から最初の'_’までの文字列を会社検索キーワードとする（最後の様などの継承は除く）。
    		アップロードしたPDFファイル内の文字列から、会社検索キーワードとして記載されている文字列を抽出してください。
    		ただし、会社検索キーワードとアルファベット、カタカナの全角、半角、大文字小文字の違い、スペースの有無は同じものとみなす。
            会社検索キーワードを返すのではなく、PDFから抽出した会社名を返すこと。
            読み取った会社名に、支社名や事業所名がある場合は、それを除いた部分を会社名とし、抽出してください。
		3)氏名
            アップロードした画像ファイル内の文字をOCRし、画像内に明確に記載された氏名のみを返してください。
            {display_file_name}の拡張子の前に数値7桁あるいは'BP'+数値5桁が社員番号である。社員番号の前の"_"から、その前の"_"までの文字列が氏名検索キーワードである。
            氏名検索キーワードはあくまで探索の補助であり、JSONの返却値ではありません。
            氏名検索キーワード、画像内の氏名の比較の際はスペース等を除いて比較してもよいが、推測で補完してはいけません。
            氏名検索キーワードを返すのではなく、画像から明確に抽出できた氏名を返してください。
            氏名検索キーワードを定義するが、これは検索結果ではない。
			アップロードした画像ファイルから氏名検索キーワードと類似した文字列を画像内から氏名として抜き出してください。
			氏名検索キーワード、画像内の氏名の比較の際、スペースなどはパックして比較して同じかどうか判断してください。
            氏名検索キーワードを返すのではなく、画像から抽出した氏名を返すこと。
            読み取れない場合は、”（データなし）”を返す
		4)年度
			年度はアップロードした画像ファイル内の勤務表から読み取れる西暦年度（4桁）を出力してください。読み取れない場合は、”（データなし）”を返す
            Mar-26などの場合は、2026年とする。
		5)月
			月はアップロードした画像ファイル内の勤務表から読み取れる月（1〜12）を出力してください。読み取れない場合は、”（データなし）”を返す
            Mar-26などの場合は、3月とする。
		6)合計勤務時間
			アップロードした画像ファイルから、総労働時間と思える文字列をそのまま抽出してください。
			例）（8:20・8時間20分・8.20・101_10H・86.17H 等。必要に応じて実データの表記のまま）
		7)押印有無
			押印有無については、以下の判断をしてください。〇、✖以外の結果は返さないでください。
			アップロードした画像ファイルの中に
  				・実際の印鑑の陰影がある：〇
  				・印鑑の陰影の印刷がある：〇
  				・サインの筆跡がある：〇
  				・抽出できない場合：✖
  				・上記以外の場合：✖
		8)交通費合計
			アップロードした画像ファイルの中で
			交通費・経費精算を部分から、交通費等の合計と思われる金額を抽出し「交通費合計」として出力してください。
			税抜き・税込みの両方の金額がある場合は、税込みの金額を採用してください。
			金額は表記のまま（円・カンマ等を含む）でよい。読み取れない場合は”（データなし）”を返す。
			読み取れない場合は、”（データなし）”を返す。
    """

# def _build_check_message(display_file_name: str) -> str:
#     """解析結果に出すファイル名をローカルの実名に固定する。"""
#     transport_block = _transport_expense_prompt_block(display_file_name)
#     transport_line = (
#         "\n交通費合計：（税抜・税込両方ある場合は税込みの金額。表記のまま）\n"
#         if transport_block
#         else ""
#     )
#     return f"""
# 勤務表の画像を解析し、１画像１行で以下の内容にあたるものをmd形式で表で出力してください。
# １）会社名の抽出について
# 画像ファイル名の会社名は会社名として出力しないこと。
# 会社名は、{display_file_name}から、先頭に[]でくくられた部分がある場合はそれ以降から次の'_’までの文字列を会社検索キーワードとしする。
# 会社名は、{display_file_name}から、先頭に[]でくくられた部分がない場合は先頭から最初の'_’までの文字列を会社検索キーワードとしする。
# 画像内から、会社検索キーワードの全部または一部をさがしてください。それを会社名として表示してください。存在しない場合は、'（存在しない）'と出力してください。
# 会社検索キーワードと読み取った会社名のアルファベット、カタカナの全角、半角、大文字小文字の違いは、同じものとみなしてください。
# 読み取った会社名にスペースを含んでいる場合もあるが、スペースは無視して、抽出してください。
# ーと-など、画像解析で読み取った文字がおおよそ類似している場合は、同一とみなして出力してください。
# 会社検索キーワードには様な度の敬称がついている場合は敬称を除外して会社検索キーワードとしてください。
# 読み取った会社名がセラクという名称を含むものは、当社の会社名なので、除外してください。
# 読み取った会社名の後ろに、支社や事業所名がある場合は、それを除いて出力してください。
# ２）氏名の抽出
# ファイル名の拡張子の前に数値7桁あるいは'BP'+数値5桁が社員番号である。その前の_から次の_までの文字列が氏名である。
# 上記の文字列を氏名として類似した文字列を画像内から氏名として抜き出してください。スペースなどがはいってる場合もあるので、無視してください。
# ３）押印有無については、以下の判断をしてください。〇、✖以外の結果は返さないでください。
#   ・実際の印鑑の陰影がある：〇
#   ・印鑑の陰影の印刷がある：〇
#   ・サインの筆跡がある：〇
#   ・抽出できない場合：✖
#   ・上記以外の場合：✖
# ４）交通費合計の抽出について
# {transport_block}
# ５）対象ファイル名について
# 画像ファイル名（アップロードファイル名）として、次の名前のみを記載してください（サーバー側のIDや別名は使わないこと）:
# {display_file_name}

# ６）出力項目は以下の通り
# 対象ファイル：{display_file_name}
# 会社名：
# 氏名：
# 年度：
# 月：
# 合計勤務時間：（8:20・8時間20分・8.20・101_10H・86.17H 等。必要に応じて実データの表記のまま）
# 押印有無：〇/✖
# {transport_line}
# """

def _build_pdf_check_message(display_file_name: str) -> str:
    return f"""
    	アップロードした画像ファイルから以下の情報を抽出してJSON形式でデータのみ返信してください。
    	
    	・対象ファイル
		・会社名
		・氏名
		・年度
		・月
		・合計勤務時間
		・押印有無
		・交通費合計

        【重要ルール】
        - 画像内に文字として明確に確認できない情報は、いかなる場合も推測・補完・類推して返さないこと。
        - 会社名、氏名は、検索キーワードに近いからという理由で補完してはならない。
        - 画像から明瞭に読み取れた文字列のみを返すこと。
        - 1文字でも判読が不確かな場合は「（データなし）」とする。
        - 画像内に存在しない、または確認できない会社名・氏名を返してはならない。

    	各項目の説明
    	1)対象ファイル
    		{display_file_name}を値として返す。
    	2)会社名
    		アップロードした画像ファイル内の文字をOCRし、画像内に明確に記載された会社名のみを返してください。
            会社検索キーワードはあくまで探索の補助であり、JSONの返却値ではありません。
            キーワードに類似していても、画像内に明確な記載が確認できない場合は「（データなし）」を返してください。
            会社検索キーワードを返すのではなく、画像から明確に抽出できた会社名を返してください。
            読み取れない、または確証がない場合は「（データなし）」を返してください。
            会社検索キーワードというものを以下に設定するがこれは返却値ではない。
    		{display_file_name}の先頭に[]でくくられた文字列がある場合、それ以降から次の'_’までの文字列を会社検索キーワードとする（最後の様などの継承は除く）。
    		{display_file_name}の先頭に[]でくくられた部分がない場合は、先頭から最初の'_’までの文字列を会社検索キーワードとする（最後の様などの継承は除く）。
    		アップロードしたPDFファイル内の文字列から、会社検索キーワードとして記載されている文字列を抽出してください。
    		ただし、会社検索キーワードとアルファベット、カタカナの全角、半角、大文字小文字の違い、スペースの有無は同じものとみなす。
            会社検索キーワードを返すのではなく、PDFから抽出した会社名を返すこと。
            読み取った会社名に、支社名や事業所名がある場合は、それを除いた部分を会社名とし、抽出してください。
		3)氏名
            アップロードした画像ファイル内の文字をOCRし、画像内に明確に記載された氏名のみを返してください。
            {display_file_name}の拡張子の前に数値7桁あるいは'BP'+数値5桁が社員番号である。社員番号の前の"_"から、その前の"_"までの文字列が氏名検索キーワードである。
            氏名検索キーワードはあくまで探索の補助であり、JSONの返却値ではありません。
            氏名検索キーワード、画像内の氏名の比較の際はスペース等を除いて比較してもよいが、推測で補完してはいけません。
            氏名検索キーワードを返すのではなく、画像から明確に抽出できた氏名を返してください。
            氏名検索キーワードを定義するが、これは検索結果ではない。
			アップロードした画像ファイルから氏名検索キーワードと類似した文字列を画像内から氏名として抜き出してください。
			氏名検索キーワード、画像内の氏名の比較の際、スペースなどはパックして比較して同じかどうか判断してください。
            氏名検索キーワードを返すのではなく、画像から抽出した氏名を返すこと。
            読み取れない場合は、”（データなし）”を返す
		4)年度
			年度はアップロードした画像ファイル内の勤務表から読み取れる西暦年度（4桁）を出力してください。読み取れない場合は、”（データなし）”を返す
            Mar-26などの場合は、2026年とする。
		5)月
			月はアップロードした画像ファイル内の勤務表から読み取れる月（1〜12）を出力してください。読み取れない場合は、”（データなし）”を返す
            Mar-26などの場合は、3月とする。
		6)合計勤務時間
			アップロードした画像ファイルから、総労働時間と思える文字列をそのまま抽出してください。
			例）（8:20・8時間20分・8.20・101_10H・86.17H 等。必要に応じて実データの表記のまま）
		7)押印有無
			押印有無については、以下の判断をしてください。〇、✖以外の結果は返さないでください。
			アップロードした画像ファイルの中に
  				・実際の印鑑の陰影がある：〇
  				・印鑑の陰影の印刷がある：〇
  				・サインの筆跡がある：〇
  				・抽出できない場合：✖
  				・上記以外の場合：✖
		8)交通費合計
			アップロードした画像ファイルの中で
			交通費・経費精算を部分から、交通費等の合計と思われる金額を抽出し「交通費合計」として出力してください。
			税抜き・税込みの両方の金額がある場合は、税込みの金額を採用してください。
			金額は表記のまま（円・カンマ等を含む）でよい。読み取れない場合は”（データなし）”を返す。
			読み取れない場合は、”（データなし）”を返す。
    """

# def _build_pdf_check_message(display_file_name: str) -> str:
#     return f"""
#     	アップロードしたPDFファイルから以下の情報を抽出してJSON形式でデータのみ返信してください。
    	
#     	・対象ファイル
# 		・会社名
# 		・氏名
# 		・年度
# 		・月
# 		・合計勤務時間
# 		・押印有無
# 		・交通費合計

#         【重要ルール】
#         - PDF内に文字として明確に確認できない情報は、いかなる場合も推測・補完・類推して返さないこと。
#         - 会社名、氏名は、検索キーワードに近いからという理由で補完してはならない。
#         - PDF内から明瞭に読み取れた文字列のみを返すこと。
#         - 1文字でも判読が不確かな場合は「（データなし）」とする。
#         - 画像内に存在しない、または確認できない会社名・氏名を返してはならない。

    	
#     	各項目の説明
#     	1)対象ファイル
#     		{display_file_name}を値として返す。
#     	2)会社名
#     		アップロードしたPDFファイル内の文字を解析し、PDF内に明確に記載された会社名のみを返してください。
#             読み取った会社名の後ろに支社や事業所名がある場合は、それを除いて出力してください。
#             会社検索キーワードを返すのではなく、PDFから明確に抽出できた会社名をJSONの結果として返してください。
#             読み取れない、または確証がない場合は「（データなし）」を返してください。
#             会社検索キーワードというものを以下に設定するがこれはJSONの返却値ではない。
#     		{display_file_name}の先頭に[]でくくられた文字列がある場合、それ以降から次の'_’までの文字列を会社検索キーワードとする（最後の様などの継承は除く）。
#     		{display_file_name}の先頭に[]でくくられた部分がない場合は、先頭から最初の'_’までの文字列を会社検索キーワードとする（最後の様などの継承は除く）。
#     		アップロードしたPDFファイル内の文字列から、会社検索キーワードとして記載されている文字列を抽出してください。
#     		ただし、会社検索キーワードとアルファベット、カタカナの全角、半角、大文字小文字の違い、スペースの有無は同じものとみなす。
#             会社検索キーワードを返すのではなく、PDFから抽出した会社名を返すこと。
#             読み取った会社名に、支社名や事業所名がある場合は、それを除いた部分を会社名とし、抽出してください。
# 		3)氏名
#             アップロードしたPDFファイル内の文字を解析し、PDF内に明確に記載された氏名のみを返してください。
#             氏名検索キーワードはあくまで探索の補助であり、返却値ではありません。
#             氏名検索キーワード、PDF内の氏名の比較の際はスペース等を除いて比較してもよいが、推測で補完してはいけません。
#             氏名検索キーワードを返すのではなく、PDFから明確に抽出できた氏名を返してください。
#             氏名検索キーワードを定義するが、これは検索結果ではない。
# 			{display_file_name}の拡張子の前に数値7桁あるいは'BP'+数値5桁が社員番号である。社員番号の前の"_"から、その前の"_"までの文字列が年月であり
#             そこからさらにその前の"_"までが氏名検索キーワードである。
# 			アップロードした画像ファイルから氏名検索キーワードと類似した文字列を画像内から氏名として抜き出してください。
# 			氏名検索キーワード、画像内の氏名の比較の際、スペースなどはパックして比較して同じかどうか判断してください。
#             氏名検索キーワードを返すのではなく、画像から抽出した氏名を返すこと。
#             読み取れない場合は、”（データなし）”を返す
# 		4)年度
# 			年度はアップロードしたPDFファイル内の勤務表から読み取れる西暦年度（4桁）を出力してください。読み取れない場合は、”（データなし）”を返す
# 		5)月
# 			月はアップロードしたPDFファイル内の勤務表から読み取れる月（1〜12）を出力してください。読み取れない場合は、”（データなし）”を返す
# 		6)合計勤務時間
# 			アップロードしたPDFファイルから、総労働時間と思える文字列をそのまま抽出してください。
# 			例）（8:20・8時間20分・8.20・101_10H・86.17H 等。必要に応じて実データの表記のまま）
# 		7)押印有無
# 			押印有無については、以下の判断をしてください。〇、✖以外の結果は返さないでください。
# 			アップロードしたPDFファイルの中に
#   				・実際の印鑑の陰影がある：〇
#   				・印鑑の陰影の印刷がある：〇
#   				・サインの筆跡がある：〇
#   				・抽出できない場合：✖
#   				・上記以外の場合：✖
# 		8)交通費合計
# 			{display_file_name}に"交通費"という文字列を含んでいる場合
# 				アップロードしたPDFファイルの中で
# 				交通費・経費精算を部分から、交通費等の合計と思われる金額を抽出し「交通費合計」として出力してください。
# 				税抜き・税込みの両方の金額がある場合は、税込みの金額を採用してください。
# 				金額は表記のまま（円・カンマ等を含む）でよい。読み取れない場合は”（データなし）”を返す。
# 			{display_file_name}に"交通費"という文字がある場合は""（空文字）を返す	
     
#     """
# def _build_pdf_check_message(display_file_name: str) -> str:
#     """PDF用。アップロード直後は send_message にドキュメントIDを渡さなくても参照できる想定。"""
#     transport_block = _transport_expense_prompt_block(display_file_name)
#     transport_pdf_note = (
#         "ファイル名に「交通費」を含むため、交通費合計は経費精算・交通費の記載から抽出してください（勤怠と経費が混在する場合、交通費合計のみ経費側を参照してよい）。\n"
#         if transport_block
#         else ""
#     )
#     return f"""
# {transport_pdf_note}
# 同じPDFに勤怠（出退勤・打刻・勤務時間等）の情報と経費精算（領収書・立替等）の情報の両方が含まれている場合は、経費精算は無視し、必ず勤怠（勤務表）の情報だけを根拠に回答してください。経費側の社名・氏名・金額は採用しないでください。
# PDFファイル名の会社名は会社名として抽出しないこと。
# 勤務表（勤怠）のPDFを解析してください。出力する勤務先・氏名・合計勤務時間・押印はすべて勤怠部分の記載に基づきます。
# １）会社名の抽出について
# 会社名は、{display_file_name}から、先頭に[]でくくられた部分がある場合はそれ以降から次の'_'までの文字列を会社検索キーワードとする。
# 会社名は、{display_file_name}から、先頭に[]でくくられた部分がない場合は先頭から最初の'_'までの文字列を会社検索キーワードとする。
# PDF内から、会社検索キーワードの全部または一部をさがしてください。存在しない場合は、'（存在しない）'と出力してください。
# ２）氏名の抽出
# ファイル名の拡張子の前に数値7桁あるいは'BP'+数値5桁が社員番号である。その前の_から次の_までの文字列が氏名である。
# 上記の文字列を氏名として類似した文字列をPDF内から氏名として抜き出してください。
# ３）押印有無については、〇、✖以外の結果は返さないでください。
# ４）交通費合計の抽出について
# {transport_block}
# ５）年度・月は勤怠表から読み取れる西暦年度（4桁）と月（1〜12）を出力してください。
# ６）合計勤務時間は勤怠部分の合計・総労働時間等を表記のまま出力してください。
# {_build_json_response_instruction(display_file_name)}
# """


def _excel_target_sheet_symbol(
    file_path: Path,
    *,
    target_sheet_name: str = TARGET_EXCEL_SHEET_NAME,
) -> tuple[str, str]:
    """Excel ブック内に対象シートがあるかを判定し、(記号, 補足メッセージ) を返す。"""
    try:
        with zipfile.ZipFile(file_path) as zf:
            with zf.open("xl/workbook.xml") as fp:
                root = ET.parse(fp).getroot()
        ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        sheet_names = [
            str(node.attrib.get("name") or "").strip()
            for node in root.findall(".//x:sheets/x:sheet", ns)
        ]
    except (FileNotFoundError, OSError, KeyError, ET.ParseError, zipfile.BadZipFile) as e:
        return "✖", f"Excelシート確認失敗: {e}"
    return ("〇", "") if target_sheet_name in sheet_names else ("✖", "")


def _excel_cell_value_to_raw_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, DTDateTime):
        return value.strftime("%H:%M")
    if isinstance(value, DTTime):
        return value.strftime("%H:%M")
    if isinstance(value, DTTimeDelta):
        total_minutes = max(0, int(round(value.total_seconds() / 60.0)))
        hh, mm = divmod(total_minutes, 60)
        return f"{hh}:{mm:02d}"
    return str(value).strip()


def _excel_cell_value_to_year(value: object) -> str:
    """Excel セル（A7 等）から年度（4桁西暦）を正規化する。"""
    if value is None:
        return ""
    if isinstance(value, DTDateTime):
        return str(value.year)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        v = int(value)
        if 1900 <= v <= 2100:
            return str(v)
    return _normalize_year_value(_excel_cell_value_to_raw_text(value))


def _excel_cell_value_to_month(value: object) -> str:
    """Excel セル（D7 等）から月（1〜12）を正規化する。"""
    if value is None:
        return ""
    if isinstance(value, DTDateTime):
        m = value.month
        return str(m) if 1 <= m <= 12 else ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        v = int(value)
        if 1 <= v <= 12:
            return str(v)
    return _normalize_month_value(_excel_cell_value_to_raw_text(value))


def _excel_cell_value_to_decimal_hours(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, DTDateTime):
        hours = (
            value.hour
            + value.minute / 60.0
            + value.second / 3600.0
            + value.microsecond / 3600000000.0
        )
        return _format_decimal_str(hours)
    if isinstance(value, DTTime):
        hours = (
            value.hour
            + value.minute / 60.0
            + value.second / 3600.0
            + value.microsecond / 3600000000.0
        )
        return _format_decimal_str(hours)
    if isinstance(value, DTTimeDelta):
        return _format_decimal_str(value.total_seconds() / 3600.0)
    if isinstance(value, (int, float)):
        return _format_decimal_str(float(value) * 24.0)
    return _work_hours_string_to_decimal(str(value))


def _employee_no_from_file_name(file_name: str) -> str:
    stem = Path(file_name or "").stem
    if len(stem) < 7:
        return ""
    tail7 = stem[-7:]
    if tail7.isdigit():
        return tail7
    if re.fullmatch(r"BP\d{5}", tail7, re.IGNORECASE):
        return tail7.upper()
    return "社員番号エラー"


def _filename_person_for_billing_lookup(person_segment: str) -> str:
    """ファイル名の会社名以降セグメントから、請求シート照合用の氏名を抽出する。"""
    skip_tokens = frozenset(
        {
            "勤務表",
            "作業報告書",
            "交通費",
            "勤務",
            "表",
            "報告書",
            "前半",
            "後半",
            "押印無",
            "押印有",
        }
    )
    parts: list[str] = []
    for p in re.split(r"[_＿]", person_segment or ""):
        t = p.strip()
        if not t or t in skip_tokens:
            continue
        if re.fullmatch(r"\d{4,}", t):
            break
        if re.fullmatch(r"BP\d{5}", t, re.IGNORECASE):
            break
        if re.fullmatch(r"\d{7}", t):
            break
        parts.append(t)
    if parts:
        return "".join(parts)
    return _normalize_person(person_segment)


def _billing_sheet_row_matches_filename(
    file_company: str,
    file_person_segment: str,
    sheet_client: str,
    sheet_person: str,
) -> bool:
    """D列取引先名・E列氏名とファイル名の会社・氏名が一致するか。"""
    fc = _company_core_for_match(file_company)
    sc = _company_core_for_match(sheet_client)
    if not fc or not sc:
        return False
    if fc not in sc and sc not in fc:
        return False
    fp = _normalize_person(_filename_person_for_billing_lookup(file_person_segment))
    sp = _normalize_person(sheet_person)
    if not fp or not sp:
        return False
    return fp == sp or (len(fp) >= 2 and (fp in sp or sp in fp))


def _load_billing_engineer_ts_rows(billing_path: Path) -> list[tuple[str, str, str]]:
    """エンジニアTS一覧の (社員番号, 取引先名, 氏名) 行リスト。"""
    path = billing_path.resolve()
    suffix = path.suffix.lower()
    keep_vba = suffix in (".xlsm", ".xltm")
    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=keep_vba)
    try:
        if BILLING_ENGINEER_TS_SHEET_NAME not in wb.sheetnames:
            return []
        ws = wb[BILLING_ENGINEER_TS_SHEET_NAME]
        rows: list[tuple[str, str, str]] = []
        max_row = ws.max_row or 0
        for row_idx in range(1, max_row + 1):
            emp = _normalize_employee_no_cell_value(
                ws.cell(row=row_idx, column=BILLING_TS_COL_EMPLOYEE_NO).value
            )
            if not _is_valid_employee_no(emp):
                continue
            client = _excel_cell_value_to_raw_text(
                ws.cell(row=row_idx, column=BILLING_TS_COL_CLIENT_NAME).value
            )
            person = _excel_cell_value_to_raw_text(
                ws.cell(row=row_idx, column=BILLING_TS_COL_PERSON_NAME).value
            )
            if not (client.strip() and person.strip()):
                continue
            rows.append((emp, client, person))
        return rows
    finally:
        wb.close()


def lookup_employee_no_in_billing_file(
    billing_path: Path,
    file_name: str,
    *,
    sheet_rows: list[tuple[str, str, str]] | None = None,
) -> str:
    """請求用ファイルのエンジニアTS一覧から、ファイル名に対応する社員番号を返す。"""
    file_co, file_pe_seg = _parse_filename_company_and_person(file_name)
    if not (file_co.strip() and file_pe_seg.strip()):
        return ""
    rows = sheet_rows if sheet_rows is not None else _load_billing_engineer_ts_rows(billing_path)
    for emp, client, person in rows:
        if _billing_sheet_row_matches_filename(file_co, file_pe_seg, client, person):
            return emp
    return ""


def _excel_transport_expense_from_sheet(ws) -> str:
    """交通宿泊費清算書シートで F列を下から走査し、「合計」行の G列を返す。"""
    max_row = ws.max_row or 0
    for row_idx in range(max_row, 0, -1):
        label = _excel_cell_value_to_raw_text(
            ws.cell(row=row_idx, column=TRANSPORT_EXPENSE_LABEL_COL).value
        )
        if label != TRANSPORT_EXPENSE_LABEL_TEXT:
            continue
        return _excel_cell_value_to_raw_text(
            ws.cell(row=row_idx, column=TRANSPORT_EXPENSE_VALUE_COL).value
        ).strip()
    return ""


def _attach_excel_transport_expense(row: dict[str, str], file_path: Path) -> None:
    """交通宿泊費清算書お客様先用 シートから交通費合計（読取）を設定する。"""
    if _filename_has_transport_expense_marker(file_path.name):
        # ファイル名に「交通費」があるのに取得できないケースは（不明）にする
        row.setdefault("transport_expense_raw", "（不明）")
    sheet_symbol, _ = _excel_target_sheet_symbol(
        file_path,
        target_sheet_name=TRANSPORT_EXPENSE_EXCEL_SHEET_NAME,
    )
    if sheet_symbol != "〇":
        return
    wb = None
    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb[TRANSPORT_EXPENSE_EXCEL_SHEET_NAME]
        row["transport_expense_raw"] = _excel_transport_expense_from_sheet(ws)
    except Exception:
        pass
    finally:
        if wb is not None:
            wb.close()


def _extract_excel_target_sheet_row(
    file_path: Path,
    *,
    target_sheet_name: str = TARGET_EXCEL_SHEET_NAME,
) -> dict[str, str]:
    row = {
        "upload_ok": "",
        "file_name": file_path.name,
        "resolved_path": str(file_path.resolve()),
        "employee_no": _employee_no_from_file_name(file_path.name),
        "target_sheet_exists": "✖",
        "analysis": "",
        "source_kind": "excel",
    }

    def _finish() -> dict[str, str]:
        _attach_excel_transport_expense(row, file_path)
        return row

    sheet_symbol, note = _excel_target_sheet_symbol(
        file_path,
        target_sheet_name=target_sheet_name,
    )
    row["target_sheet_exists"] = sheet_symbol
    if note:
        row["analysis"] = note
        return _finish()

    # 1) 通常: TARGET_EXCEL_SHEET_NAME があれば従来通り読み取る
    if sheet_symbol == "〇":
        wb = None
        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            ws = wb[target_sheet_name]
            company = (ws["G5"].value or "")
            person = (ws["J7"].value or "")
            total_raw_value = ws["F42"].value
            doc_y = _excel_cell_value_to_year(ws["A7"].value)
            doc_m = _excel_cell_value_to_month(ws["D7"].value)
            if doc_y:
                row["year"] = doc_y
            if doc_m:
                row["month"] = doc_m

            file_company, _file_person = _parse_filename_company_and_person(file_path.name)
            doc_company_raw = str(company).strip() if company is not None else ""
            doc_company_match = _document_company_for_match(doc_company_raw)
            match_company = _match_company_symbol_single(file_company, doc_company_match)

            row["name_company_from_file"] = file_company
            row["name_company_from_doc"] = doc_company_raw
            row["name_company_1"] = str(company).strip() if company is not None else ""
            row["name_person_from_doc"] = str(person).strip() if person is not None else ""
            row["total_hours_raw"] = _excel_cell_value_to_raw_text(total_raw_value)
            row["total_hours_decimal"] = _excel_cell_value_to_decimal_hours(total_raw_value)
            row["match_company"] = match_company
            # 仕様: 1行解析時は自動判断をユーザ判断にもセットする
            aj = _auto_judgment_symbol(row)
            row["auto_judgment"] = aj
            row["user_judgment_company"] = aj
        except Exception as e:
            row["analysis"] = f"Excelセル読み取り失敗: {e}"
        finally:
            if wb is not None:
                wb.close()
        return _finish()

    # 対象シートなし: 対象シート有無 ✖ のまま終了
    return _finish()


# --- 合計勤務時間の10進化 -----------------------------------------------------------------
def _trunc2_down(x: float) -> float:
    """正の数について、小数点以下第3位以下を切り捨て、第2位まで有効化する（100倍floor/100）。"""
    n = 100
    return math.floor((x if x >= 0 else 0) * n + 1e-9) / n


def _format_decimal_str(value: float) -> str:
    """小数点以下第3位以下切り捨てのうえ、小数第2位まで保持（整数は .00 なし）。"""
    t = _trunc2_down(value)
    if math.isclose(t, round(t), abs_tol=1e-9):
        return str(int(round(t)))
    return f"{t:.2f}"


def _work_hours_raw_to_hours_float(raw: str) -> float | None:
    """合計勤務時間（読取）文字列を10進時間（float）に変換。切り捨て前の値を返す。"""
    if not (raw and str(raw).strip()):
        return None
    t = unicodedata.normalize("NFKC", str(raw).strip())
    t = t.split("(")[0].split("（")[0]
    t = re.sub(r"[\s　]+", "", t)
    t = t.replace("：", ":").replace("．", ".")
    if not t:
        return None

    m = re.match(r"^(\d+)_(\d{1,2})[Hh]$", t)
    if m:
        try:
            hh, mm = int(m.group(1)), int(m.group(2))
            if mm > 59:
                return None
            return float(hh) + mm / 60.0
        except (ValueError, OverflowError):
            return None

    m = re.match(r"^(\d+):(\d{1,2})$", t)
    if m:
        try:
            hh, mm = int(m.group(1)), int(m.group(2))
            if mm > 59:
                return None
            return float(hh) + mm / 60.0
        except (ValueError, OverflowError):
            return None

    m = re.match(r"^(\d+):(\d{1,2}):(\d{1,2})$", t)
    if m:
        try:
            hh, mm, ss = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if mm > 59 or ss > 59:
                return None
            return float(hh) + mm / 60.0
        except (ValueError, OverflowError):
            return None

    m = re.match(r"^(\d+)時間(\d{1,2})分$", t)
    if m:
        try:
            hh, mm = int(m.group(1)), int(m.group(2))
            if mm > 59:
                return None
            return float(hh) + mm / 60.0
        except (ValueError, OverflowError):
            return None

    m = re.match(r"^(\d+)時間$", t)
    if m:
        try:
            return float(int(m.group(1)))
        except (ValueError, OverflowError):
            return None

    m = re.match(r"^(\d+\.\d+)[Hh]$", t)
    if m:
        try:
            return float(m.group(1))
        except (ValueError, OverflowError):
            return None

    m = re.match(r"^(\d+)\.(\d+)$", t)
    if m:
        try:
            h, fr = m.group(1), m.group(2)
            return int(h) + int(fr) / 10.0 ** len(fr)
        except (ValueError, OverflowError):
            return None

    m = re.match(r"^(\d+)[Hh]$", t)
    if m:
        try:
            return float(m.group(1))
        except (ValueError, OverflowError):
            return None

    m = re.match(r"^(\d+)$", t)
    if m:
        try:
            return float(m.group(1))
        except (ValueError, OverflowError):
            return None

    return None


def _work_hours_string_to_decimal(raw: str) -> str:
    """
    合計（総）勤務時間の10進2桁化（第3位未満切り捨て）。
    - 8.35  … 十進（小数点はそのまま時間。8.35 時間 = 8.35）
    - 8.35H … 十進（N.NH は N.N 時間）
    - 8:35  … 60進 → H+MM/60
    - 144:00:00 … 60進（秒は切り捨て、時・分のみ換算）
    - 8時間35分 … 60進
    - 8_35H  … 60進（時_分H）
    - 140   … 十進（時間のみの整数。140 時間 = 140）
    - 140H  … 十進（整数 + H はそのまま時間）
    解釈できなければ空文字。
    """
    val = _work_hours_raw_to_hours_float(raw)
    if val is None:
        return ""
    return _format_decimal_str(val)


def _md_table_row_cells(line: str) -> list[str]:
    s = (line or "").strip()
    if not s.startswith("|"):
        return []
    inner = s[1:-1] if s.endswith("|") and len(s) > 1 else s[1:]
    return [c.strip() for c in inner.split("|")]


def _is_md_table_separator_row(cells: list[str]) -> bool:
    if not cells:
        return True
    for c in cells:
        t = c.replace(" ", "").replace("\u3000", "")
        if re.search(r"[^:｜\-\s─━]", t):
            return False
    return True


# 区切り行の隣接セルに誤ってマッチしないよう、列名そのものを弾く
_KINTAI_HEADER_CELLS: frozenset[str] = frozenset(
    {
        "会社名",
        "勤務先",
        SUMMARY_PERSON_COL,
        LEGACY_PERSON_COL,
        "氏名",
        SUMMARY_EMPLOYEE_NO_COL,
        LEGACY_EMPLOYEE_NO_COL,
        "社員番号",
        "合計勤務時間",
        "合計",
        "押印有無",
        "押印",
        "有無",
        TARGET_FILE_NAME_COL,
        LEGACY_TARGET_FILE_NAME_COL,
        LEGACY_FILE_NAME_COL,
        SUMMARY_COMPANY_COL,
        LEGACY_COMPANY_READ_LONG_COL,
        LEGACY_COMPANY_COL,
        LEGACY_COMPANY_NAME_COL,
        "ファイル名",
        "年度",
        SUMMARY_YEAR_COL,
        SUMMARY_MONTH_COL,
        LEGACY_YEAR_COL,
        LEGACY_MONTH_COL,
        "年",
        "月",
        "交通費合計",
        "交通費",
    }
)


def _is_table_headerish_cell(s: str) -> bool:
    t = unicodedata.normalize("NFKC", (s or "").strip())
    if t in _KINTAI_HEADER_CELLS or t in ("画像", "合計勤務", "労働時間"):
        return True
    if re.fullmatch(r"氏名[（(].+?[)）]?", t):
        return True
    return bool(re.match(r"^(合計|押印|勤務先|画像|ファイル)", t) and len(t) <= 20)


def _kintai_header_col_roles(cells: list[str]) -> dict[str, int] | None:
    """
    見出し行: 各列の役 (co, pe, th, se, fn) → 0始まりインデックス
    勤怠表らしい列が2つ以上特定できたときだけ返す。
    """
    roles: dict[str, int] = {}
    for j, c in enumerate(cells):
        t = unicodedata.normalize("NFKC", (c or "").strip())
        if not t:
            continue
        if "co" not in roles and (
            t in ("会社名", "勤務先", "勤務先会社", "就業先")
            or ("勤務" in t and "会社" in t)
        ):
            roles["co"] = j
        elif "pe" not in roles and (t == "氏名" or t.startswith("氏名")):
            roles["pe"] = j
        elif "th" not in roles and (
            t in ("合計勤務時間", "合計勤務", "総労働時間", "労働時間", "合計", "労働")
            or (("合計" in t or "総" in t) and ("勤務" in t or "時間" in t or "労働" in t))
        ):
            roles["th"] = j
        elif "se" not in roles and (t in ("押印有無",) or ("押印" in t) or t == "有無"):
            roles["se"] = j
        elif "te" not in roles and (
            t in ("交通費合計", "交通費") or "交通費合計" in t
        ):
            roles["te"] = j
        elif "fn" not in roles and (
            t in (
                TARGET_FILE_NAME_COL,
                LEGACY_TARGET_FILE_NAME_COL,
                LEGACY_FILE_NAME_COL,
                "ファイル名",
            )
            or (("画像" in t or "対象" in t) and "ファイル" in t)
        ):
            roles["fn"] = j
        elif "yr" not in roles and t in ("年度", "年"):
            roles["yr"] = j
        elif "mo" not in roles and t == "月":
            roles["mo"] = j
    if "co" in roles and "pe" in roles:
        return roles
    if "pe" in roles and "th" in roles:
        return roles
    if "co" in roles and "th" in roles:
        return roles
    if "fn" in roles and (len(roles) >= 2):
        return roles
    if len(roles) >= 3:
        return roles
    return None


def _extract_kintai_from_markdown_table(
    text: str, expected_file_name: str
) -> dict[str, str] | None:
    """
    LLMの |...| 表から、見出し行の次以降のデータ行のセルを取る（隣接セルに列名が入る行はスキップ）。
    """
    if not (text and str(text).strip()):
        return None
    exp = Path(expected_file_name or "").name

    table_lines: list[list[str]] = []
    for _i, line in enumerate(str(text).splitlines()):
        s = line.strip()
        if not s.startswith("|"):
            continue
        cells = _md_table_row_cells(s)
        if not cells:
            continue
        if _is_md_table_separator_row(cells):
            continue
        table_lines.append(cells)

    def _row_looks_header_only(c: list[str]) -> bool:
        nonempty = [x for x in c if (x or "").strip()]
        if not nonempty:
            return True
        return all(_is_table_headerish_cell(x) for x in nonempty)

    def _cell_get(cells: list[str], key: str, roles: dict[str, int]) -> str:
        if key not in roles:
            return ""
        j = roles[key]
        if j < 0 or j >= len(cells):
            return ""
        v = (cells[j] or "").strip()
        if not v or _is_table_headerish_cell(v):
            return ""
        return v

    def _build_row(dcells: list[str], roles: dict[str, int]) -> dict[str, str]:
        return {
            "company": _cell_get(dcells, "co", roles),
            "person": _cell_get(dcells, "pe", roles),
            "total": _cell_get(dcells, "th", roles),
            "seal": _cell_get(dcells, "se", roles),
            "year": _cell_get(dcells, "yr", roles),
            "month": _cell_get(dcells, "mo", roles),
            "transport": _cell_get(dcells, "te", roles),
        }

    def _row_matches_filename(
        dcells: list[str], roles: dict[str, int]
    ) -> bool:
        if "fn" not in roles or not exp:
            return True
        vfn = _cell_get(dcells, "fn", roles)
        if not vfn:
            return True
        vfnn = unicodedata.normalize("NFKC", vfn)
        st = Path(exp).stem
        stn = unicodedata.normalize("NFKC", st)
        return not (
            st not in vfnn
            and stn not in vfnn
            and exp not in vfnn
            and vfnn not in exp
        )

    for require_fn in (True, False):
        for hi, hcells in enumerate(table_lines):
            roles = _kintai_header_col_roles(hcells)
            if not roles:
                continue
            for sub in range(hi + 1, len(table_lines)):
                dcells = table_lines[sub]
                if _kintai_header_col_roles(dcells) is not None:
                    break
                if not dcells or _row_looks_header_only(dcells):
                    continue
                if require_fn and not _row_matches_filename(dcells, roles):
                    continue
                d = _build_row(dcells, roles)
                if any(d.values()):
                    return d
    return None


def _is_plausible_total_labor_value(v: str) -> bool:
    """表のヘッダ行の隣セル（列名等）を拾わない: 数値必須。"""
    t = (v or "").strip()
    if not t or len(t) > 120:
        return False
    if t in ("-", "—", "未記載", "n/a", "N/A", "null", "要確認", "（なし）"):
        return False
    if _is_table_headerish_cell(t) and not re.search(r"\d", t):
        return False
    return bool(re.search(r"\d", t))


def _extract_total_work_hours_from_document(text: str) -> str:
    """解析文から 合計勤務時間 ラベルに続く1トークンを抜出（文面上の先頭候補で数値含むもののみ）。"""
    if not (text and str(text).strip()):
        return ""
    raw = str(text)
    cands: list[tuple[int, str]] = []
    _skip = ("", "-", "—", "未記載", "n/a", "N/A", "null", "要確認")
    for m in re.finditer(
        r"合計勤務時間\s*[：:｜\|]?\s*([^\n\r\|]+?)(?=(\s*\||$|\n))",
        raw,
    ):
        v = m.group(1).strip()
        if v and v not in _skip:
            cands.append((m.start(1), v))
    for m in re.finditer(
        r"\|[^\n]*?合計勤務時間(?:\s*\|)?\s*([^\n\|]+)", raw, re.IGNORECASE
    ):
        v = m.group(1).strip()
        if v and v not in _skip:
            cands.append((m.start(1), v))
    cands.sort(key=lambda x: (x[0], x[1]))
    for _pos, v in cands:
        if _is_plausible_total_labor_value(v):
            return v
    return ""


def _pick_transport_expense_amount_display(fragment: str) -> str:
    """金額断片から表示用文字列を選ぶ（税込みを税抜きより優先）。"""
    if not (fragment and str(fragment).strip()):
        return ""
    t = unicodedata.normalize("NFKC", str(fragment).strip())
    _skip = {"", "-", "—", "未記載", "n/a", "N/A", "null", "要確認", "（なし）", "不明"}
    if t in _skip:
        return ""
    parts: list[tuple[int, str]] = []
    for m in re.finditer(
        r"(税込み?|税抜き?|税抜)?\s*([\d,，、]+(?:\.\d+)?)\s*(円)?",
        t,
    ):
        label = (m.group(1) or "").strip()
        num = (m.group(2) or "").strip()
        yen = (m.group(3) or "").strip()
        if not num or not re.search(r"\d", num):
            continue
        disp = f"{num}{yen}" if yen else num
        pri = 2 if "税込" in label else (1 if "税抜" in label else 0)
        parts.append((pri, disp))
    if parts:
        parts.sort(key=lambda x: x[0], reverse=True)
        if parts[0][0] >= 2:
            return parts[0][1]
        if len(parts) == 1:
            return parts[0][1]
    return t


def _extract_transport_expense_from_document(text: str) -> str:
    """解析文から 交通費合計 ラベルに続く金額を抜出（税込み優先）。"""
    if not (text and str(text).strip()):
        return ""
    raw = str(text)
    _skip = ("", "-", "—", "未記載", "n/a", "N/A", "null", "要確認")
    cands: list[tuple[int, str]] = []
    for m in re.finditer(
        r"交通費合計\s*[：:｜\|]?\s*([^\n\r\|]+?)(?=(\s*\||$|\n))",
        raw,
    ):
        v = m.group(1).strip()
        if v and v not in _skip:
            cands.append((m.start(1), v))
    for m in re.finditer(
        r"\|[^\n]*?交通費合計(?:\s*\|)?\s*([^\n\|]+)", raw, re.IGNORECASE
    ):
        v = m.group(1).strip()
        if v and v not in _skip:
            cands.append((m.start(1), v))
    if not cands:
        for m in re.finditer(
            r"交通費[^\n]{0,120}?合計\s*[：:｜\|]?\s*([^\n\r\|]+?)(?=(\s*\||$|\n))",
            raw,
        ):
            v = m.group(1).strip()
            if v and v not in _skip:
                cands.append((m.start(1), v))
    if not cands:
        return ""
    cands.sort(key=lambda x: (x[0], x[1]))
    picked = _pick_transport_expense_amount_display(cands[0][1])
    return picked or cands[0][1]


def _decimal_for_table_display(s: str) -> str:
    """10進列: 変換結果を表示（不要な末尾 .00 は落とし、8.35 / 98.5 のようになる）。"""
    t = (s or "").strip()
    if not t:
        return "（なし）"
    try:
        x = float(t)
    except ValueError:
        return t
    a = f"{x:.2f}"
    return a.rstrip("0").rstrip(".")


# --- ファイル名と解析文字列の 会社名・氏名 照合（〇／△／✖）----------------------------------------
# 法人格表記（文書内比較前に反復除去）
_LEGAL_TOKENS = (
    "株式会社", "（株）", "㈱", "有限会社", "（有）", "合名会社", "合資会社", "合同会社",
    "一般社団法人", "一般財団法人", "ＮＰＯ法人", "NPO法人", "医療法人", "学校法人", "宗教法人",
    "独立行政法人", "(株)", "（有)", "(有)",
)
_LEGAL_TOKENS_SORTED = tuple(sorted(_LEGAL_TOKENS, key=len, reverse=True))

# 系列・グループ名寄せ用キーワード
_GROUP_MARKERS = (
    "ホールディング", "ホールディングス", "グループ",
)


def _strip_trailing_sama(company_segment: str) -> str:
    """末尾の「様」を除去。ファイル名の会社部分に加え、照合用に読み取り側の会社名にも使う。"""
    t = unicodedata.normalize("NFKC", (company_segment or "").strip())
    if t.endswith("様"):
        t = t[: -1].rstrip(" 　\t")
    return t


def _parse_filename_company_and_person(file_name: str) -> tuple[str, str]:
    """ファイル名（拡張子除く）から会社名と氏名を分離。

    先頭 […] を繰り返し除去し、会社側の末尾「様」を除去。

    例:
      - [xxx]AAA様_山田太郎_123456.jpg -> (AAA, 山田太郎)
      - AAA様 山田太郎 123456.png -> (AAA, 山田太郎)

    ※社員番号は別ロジックで抽出して表示に付与する。
    """
    stem = Path(file_name).stem
    s = stem
    while True:
        m = re.match(r"^\[([^\]]+)\]\s*", s)
        if m:
            s = s[m.end() :]
        else:
            break
    s = s.strip()

    # 後ろに社員番号等が付いていても、会社名/氏名の分割は「最初の区切り」で行う
    for sep in ("_", "＿", "·", "・", " ", "　"):
        if sep in s:
            a, b = s.split(sep, 1)
            return _strip_trailing_sama(a), b.strip()
    return _strip_trailing_sama(s), ""


def _strip_leading_legal_tokens(t: str, tokens: tuple[str, ...]) -> str:
    """先頭に連続する法人格・組織形態を長い表記から順に剥がす。"""
    t = t.lstrip(" 　、，・.")
    while t:
        stripped = False
        for w in tokens:
            if t.startswith(w):
                t = t[len(w) :].lstrip(" 　、，・.")
                stripped = True
                break
        if not stripped:
            break
    return t


def _strip_legal_forms(company: str) -> str:
    """法人格・組織形態を先頭・文中・末尾のいずれでも除去（反復。長い表記を優先）。"""
    t = unicodedata.normalize("NFKC", str(company).strip())
    for _ in range(14):
        t0 = t
        t = _strip_leading_legal_tokens(t, _LEGAL_TOKENS_SORTED)
        for w in _LEGAL_TOKENS_SORTED:
            t = t.replace(w, "")
        t = re.sub(r"[\s　、，・\.]+", "", t)
        if t == t0:
            break
    return t


def _normalize_mixed(s: str) -> str:
    """数値・英字: 全半角・大文字小文字のゆらぎを吸収。"""
    t = unicodedata.normalize("NFKC", s or "")
    out: list[str] = []
    for c in t:
        if c.isascii() and c.isalpha():
            out.append(c.lower())
        else:
            out.append(c)
    return "".join(out)


def _normalize_person(s: str) -> str:
    return _normalize_mixed(re.sub(r"[\s\u3000\u00a0]+", "", s or ""))


def _company_core_for_match(s: str) -> str:
    """拠点語尾を外したコア比較用文字列。様→法人格除去の順（読取・ファイル名とも）。"""
    t = _strip_trailing_sama(s)
    t = _strip_legal_forms(t)
    t = _normalize_mixed(t)
    t = re.sub(
        r"(本社|支社|支店|営業所|工場|事業所|オフィス|店|エリア|地区)$", "", t
    )
    return t


def _prefer_kintai_text_for_extraction(text: str) -> str:
    """
    解析テキストに勤怠と経費精算の両方が混在する想定のとき、照合・勤務先・氏名の抽出元を勤怠側に寄せる。
    （LLMが両方出した場合のフォロー。見出し・先頭出現位置のヒューリスティクス。）
    """
    t = str(text) if text else ""
    if not t.strip():
        return t
    has_kin = any(
        k in t
        for k in (
            "勤怠",
            "勤務表",
            "出退勤",
            "出退",
            "合計勤務",
            "合計勤務時間",
            "打刻",
            "会社名",
            "押印",
        )
    )
    has_kei = "経費" in t or "精算" in t
    if not (has_kin and has_kei):
        return t
    keihi_pos: int | None = None
    for m in re.finditer(
        r"(?:^|\n)#{1,3}\s*[^\n]*?(?:経費精算|精算[（(]?経費|経費[（(]?精算|領収|立替|交通費?)",
        t,
        re.MULTILINE,
    ):
        p = m.start()
        if keihi_pos is None or p < keihi_pos:
            keihi_pos = p
    if keihi_pos is None:
        p0 = t.find("経費精算")
        if p0 >= 0:
            keihi_pos = p0
    if keihi_pos is None:
        return t
    head = t[:keihi_pos]
    if any(
        k in head
        for k in (
            "勤怠",
            "勤務表",
            "出退",
            "合計勤務",
            "合計勤務時間",
            "打刻",
            "勤務先",
        )
    ):
        return head.rstrip()
    # 文書上で経費が先・勤怠が後の可能性
    tail = t[keihi_pos + 1 :]
    for kw in ("勤怠", "勤務表", "出退勤", "出退", "合計勤務", "合計勤務時間", "会社名"):
        i = tail.find(kw)
        if i >= 0:
            return tail[i:].strip()
    return t


def _katakana_latin_mismatch_both_look_like_transcription(a: str, b: str) -> bool:
    """片方が主にカタカナ、他方が主にラテン字なら 読み揺れ △ 候補とみなす。"""
    def mostly_kat(x: str) -> bool:
        u = re.sub(r"[\s\w\-+]", "", x)
        if len(u) < 2:
            return False
        return not re.search(r"[^\u30a0-\u30ff\u30f4]", u)

    def mostly_lat(x: str) -> bool:
        return bool(re.search(r"[A-Za-z]{2,}", _normalize_mixed(x)))

    return (mostly_kat(a) and mostly_lat(b)) or (mostly_kat(b) and mostly_lat(a))


def _company_text_contains_seraku(s: str) -> bool:
    """社名文字列にセラクまたは英字 seraku（大小無視）が含まれるか。"""
    if not (s or "").strip():
        return False
    t = re.sub(r"\s+", "", unicodedata.normalize("NFKC", s))
    return bool(re.search(r"セラク|seraku", t, re.IGNORECASE))


def _file_co_for_match(company_from_file: str) -> str:
    """ファイル名由来の会社。セラク/seraku を含むときは照合対象外（空文字）。"""
    s = (company_from_file or "").strip()
    if not s or _company_text_contains_seraku(s):
        return ""
    return s


def _document_company_for_display(ktab_company: str) -> str:
    """会社名列用。セラクを含む勤怠表の会社は表示しない（不明扱い）。"""
    k = (ktab_company or "").strip()
    if not k:
        return "不明"
    if _company_text_contains_seraku(k):
        return "不明"
    return k


def _document_company_for_match(ktab_company: str) -> str:
    """照合用の文書側会社。複数候補はカンマ区切りのまま返し、比較側で分解する。"""
    return (ktab_company or "").strip()


def _split_document_company_candidates(companies_text: str) -> list[str]:
    """会社名の複数候補を分解する。カンマ区切り想定だが、日本語読点等も緩く許容する。"""
    raw = unicodedata.normalize("NFKC", (companies_text or "").strip())
    if not raw:
        return []
    parts = re.split(r"\s*[,，、]\s*", raw)
    out: list[str] = []
    for p in parts:
        t = (p or "").strip()
        if not t:
            continue
        if _company_text_contains_seraku(t):
            continue
        out.append(t)
    return out


def _match_company_symbol_single(file_co: str, doc_company: str) -> str:
    """勤怠表から取れた会社名とファイル名会社を照合。複数社名は候補のいずれか一致で判定。"""
    fp = _file_co_for_match(file_co)
    doc_candidates = _split_document_company_candidates(doc_company)
    if not doc_candidates:
        return "〇" if not fp else "✖"
    if not fp:
        return "△"
    symbols = [_compare_company(fp, d) for d in doc_candidates]
    if "〇" in symbols:
        return "〇"
    if "△" in symbols:
        return "△"
    return "✖"


def _series_company_hint(f_raw: str, d_raw: str) -> bool:
    """系列・ホール等の表記違いを粗く △ 相当とする。"""
    c1, c2 = _company_core_for_match(f_raw), _company_core_for_match(d_raw)
    if c1 and c1 == c2:
        return False
    for g in _GROUP_MARKERS:
        f_has, d_has = g in f_raw, g in d_raw
        if f_has != d_has and (c1 in c2 or c2 in c1):
            return True
    return False



def _normalize_year_value(raw: str) -> str:
    """年度表記を4桁の西暦年に正規化する。"""
    t = unicodedata.normalize("NFKC", (raw or "").strip())
    if not t or _is_table_headerish_cell(t):
        return ""
    m = re.search(r"(20\d{2}|19\d{2})", t)
    return m.group(1) if m else ""


def _normalize_month_value(raw: str) -> str:
    """月表記を 1〜12 の整数文字列に正規化する。"""
    t = unicodedata.normalize("NFKC", (raw or "").strip())
    if not t or _is_table_headerish_cell(t):
        return ""
    t = re.sub(r"月\s*$", "", t).strip()
    m = re.fullmatch(r"(\d{1,2})", t)
    if not m:
        return ""
    v = int(m.group(1))
    return str(v) if 1 <= v <= 12 else ""


def _extract_year_from_document(text: str) -> str:
    """年度を文書（ラベル行・Markdown表）から抽出する。"""
    if not (text and str(text).strip()):
        return ""
    raw = str(text)
    cands: list[tuple[int, str]] = []
    for m in re.finditer(
        r"年度\s*[：:｜\|]?\s*([^\n\r\|]+?)(?=(\s*\||$|\n))", raw
    ):
        v = _normalize_year_value(m.group(1))
        if v:
            cands.append((m.start(1), v))
    for m in re.finditer(
        r"\|[^\n]*?年度(?:\s*\|)?\s*([^\n\|]+)", raw, re.IGNORECASE
    ):
        v = _normalize_year_value(m.group(1))
        if v:
            cands.append((m.start(1), v))
    cands.sort(key=lambda x: (x[0], x[1]))
    return cands[0][1] if cands else ""


def _extract_month_from_document(text: str) -> str:
    """月を文書（ラベル行・Markdown表）から抽出する。"""
    if not (text and str(text).strip()):
        return ""
    raw = str(text)
    cands: list[tuple[int, str]] = []
    for m in re.finditer(
        r"(?<![年度])月\s*[：:｜\|]?\s*([^\n\r\|]+?)(?=(\s*\||$|\n))", raw
    ):
        v = _normalize_month_value(m.group(1))
        if v:
            cands.append((m.start(1), v))
    for m in re.finditer(
        r"\|[^\n]*?(?<![年度])月(?:\s*\|)?\s*([^\n\|]+)", raw, re.IGNORECASE
    ):
        v = _normalize_month_value(m.group(1))
        if v:
            cands.append((m.start(1), v))
    cands.sort(key=lambda x: (x[0], x[1]))
    return cands[0][1] if cands else ""


def _year_month_matches_expected(row: dict[str, str]) -> bool:
    """読み取った年・月が画面上の期待値（コンボ）と一致するか。"""
    exp_y = _normalize_year_value(row.get("expected_year") or "")
    exp_m = _normalize_month_value(row.get("expected_month") or "")
    if not exp_y or not exp_m:
        return False
    got_y = _normalize_year_value(
        row.get("year")
        or row.get(SUMMARY_YEAR_COL)
        or row.get(LEGACY_YEAR_COL)
        or ""
    )
    got_m = _normalize_month_value(
        row.get("month")
        or row.get(SUMMARY_MONTH_COL)
        or row.get(LEGACY_MONTH_COL)
        or ""
    )
    if not got_y or not got_m:
        return False
    return got_y == exp_y and got_m == exp_m


def _extract_person_name_from_document(text: str) -> str:
    """氏名を文から抽出。"""
    if not (text and str(text).strip()):
        return ""
    raw = str(text)
    cands: list[tuple[int, str]] = []
    for m in re.finditer(
        r"氏名\s*[：:｜\|]?\s*([^\n\r\|]+?)(?=(\s*\||$|\n))", raw
    ):
        v = m.group(1).strip()
        if v and v not in ("", "-", "—", "未記載") and not _is_table_headerish_cell(v):
            cands.append((m.start(1), v))
    for m in re.finditer(
        r"\|[^\n]*?氏名(?:\s*\|)?\s*([^\n\|]+)", raw, re.IGNORECASE
    ):
        v = m.group(1).strip()
        if v and not _is_table_headerish_cell(v):
            cands.append((m.start(1), v))
    cands.sort(key=lambda x: (x[0], x[1]))
    return cands[0][1] if cands else ""


def _normalize_seal_phrase(s: str) -> str:
    """押印欄の短い文を 〇/✖ 等に寄せる。"""
    s = (s or "").strip()
    if not s:
        return ""
    # 見出しセル（「押印有無」等）を弾く。ただし記号自体は許可。
    if _is_table_headerish_cell(s) and s not in ("有", "無", "有無", "〇", "✖", "○", "×", "✕"):
        return ""

    t = unicodedata.normalize("NFKC", s).replace(" ", "")

    # 否定系 → ✖
    if re.search(r"^(無|不|否|未|な(い|し)?|印な|印の記載な|－|—|N/?A)", t):
        return "✖"
    # 肯定系 → 〇
    if re.search(r"^(有|是|印あり?|有印|○|〇)", t):
        return "〇"

    # 文中に含まれるケース
    if "無" in t and "有" not in t:
        return "✖"
    if "有" in t and "無" not in t:
        return "〇"

    # 既に記号のとき（× は NFKC で × になる）
    if t in ("×", "✕", "✖"):
        return "✖"
    if t in ("○", "〇"):
        return "〇"

    return s[:10]


def _extract_seal_in_from_document(text: str) -> str:
    """押印有無: 有 / 無 等を文から抜出（表記は短く揃える）。"""
    if not (text and str(text).strip()):
        return ""
    raw = str(text)
    cands: list[str] = []
    for m in re.finditer(
        r"押印有無\s*[：:｜\|]?\s*([^\n\r\|]+?)(?=(\s*\||$|\n))", raw
    ):
        v = m.group(1).strip()
        if v and v not in ("", "-", "—", "未記載", "N/A", "n/a", "null", "要確認"):
            if not _is_table_headerish_cell(v) or v in ("有", "無"):
                cands.append(v)
    for m in re.finditer(
        r"\|[^\n]*?押印有無(?:\s*\|)?\s*([^\n\|]+)", raw, re.IGNORECASE
    ):
        v = m.group(1).strip()
        if v and v not in ("", "-", "—") and (
            not _is_table_headerish_cell(v) or v in ("有", "無")
        ):
            cands.append(v)
    s = (cands[0] if cands else "").strip()
    if not s:
        return ""
    return _normalize_seal_phrase(s)


def _compare_person(filename_person: str, document_person: str) -> str:
    """氏名: 姓名間の空白は正規化して比較。一致→〇 部分一致→△ それ以外✖。"""
    f = (filename_person or "").strip()
    d = (document_person or "").strip()
    if not f or not d:
        return "✖"
    a, b = _normalize_person(f), _normalize_person(d)
    if a == b:
        return "〇"
    if a in b or b in a:
        return "△"
    return "✖"


def _compare_company(
    company_from_file: str,
    company_from_doc: str,
) -> str:
    """
    ファイル名の会社と文書の会社名を比較（_company_core_for_match で双方とも末尾 様・法人格を除去）。
    コアが一致→〇。
    コアが完全一致でなくても、一方が他方の部分文字列なら△（短い方2文字以上）。
    そのほか読み揺れ・系列風は△。それ以外✖。
    """
    f = (company_from_file or "").strip()
    d = (company_from_doc or "").strip()
    if not f or not d:
        return "✖"
    cfn = _company_core_for_match(f)
    cdoc = _company_core_for_match(d)
    if not cfn or not cdoc:
        return "✖"
    if cfn == cdoc:
        return "〇"
    shorter, longer = (cfn, cdoc) if len(cfn) <= len(cdoc) else (cdoc, cfn)
    if len(shorter) >= 2 and shorter in longer:
        return "△"
    if _katakana_latin_mismatch_both_look_like_transcription(cfn, cdoc):
        return "△"
    if _series_company_hint(f, d):
        return "△"
    return "✖"


def _enrich_with_match_scores(
    row: dict[str, str],
    analysis: str,
    *,
    prefer_kintai_section: bool = False,
    analysis_json: dict[str, str] | None = None,
) -> None:
    """比較記号（会社名・氏名）と参照文字列を row に追加。

    追加仕様:
      - 画像ファイル名の「拡張子の前にある7桁」を社員番号として扱い、氏名の右（別列）に表示する。
      - 末尾7文字が存在するのに数字7桁ではない場合、「社員番号エラー」と表示する。
      - analysis_json 指定時は Markdown 表ではなく JSON 応答を優先する。
    """

    fn = row.get("file_name", "")
    if analysis_json:
        ktab = _analysis_json_to_kintai_tab(analysis_json)
        text_for_extraction = analysis
    else:
        text_for_extraction = (
            _prefer_kintai_text_for_extraction(analysis)
            if prefer_kintai_section
            else analysis
        )
        ktab = _extract_kintai_from_markdown_table(text_for_extraction, fn) or {}
    fp_co, fp_pe = _parse_filename_company_and_person(fn)
    row["employee_no"] = _employee_no_from_file_name(fn)
    k_co = (ktab.get("company") or "").strip()
    name_company_1 = _document_company_for_display(k_co)
    doc_co_match = _document_company_for_match(k_co)
    d_pe = (ktab.get("person") or "").strip() or _extract_person_name_from_document(
        text_for_extraction
    )
    th_raw = (ktab.get("total") or "").strip() or _extract_total_work_hours_from_document(
        text_for_extraction
    )
    s_seal = (ktab.get("seal") or "").strip()
    _seal_norm = _normalize_seal_phrase(s_seal) if s_seal else ""
    doc_y_raw = (ktab.get("year") or "").strip() or _extract_year_from_document(
        text_for_extraction
    )
    doc_m_raw = (ktab.get("month") or "").strip() or _extract_month_from_document(
        text_for_extraction
    )
    doc_y = _normalize_year_value(doc_y_raw) or doc_y_raw
    doc_m = _normalize_month_value(doc_m_raw) or doc_m_raw
    if doc_y:
        row["year"] = doc_y
    if doc_m:
        row["month"] = doc_m
    row["name_company_from_file"] = fp_co
    row["name_person_from_file"] = fp_pe
    row["name_company_1"] = name_company_1
    # 互換: 従来キーも残す（勤怠表の会社セルのみ）
    row["name_company_from_doc"] = k_co
    row["name_person_from_doc"] = d_pe
    row["match_company"] = _match_company_symbol_single(fp_co, doc_co_match)
    row["match_person"] = _compare_person(fp_pe, d_pe)
    th_dec = _work_hours_string_to_decimal(th_raw)
    row["total_hours_raw"] = th_raw
    row["total_hours_decimal"] = th_dec
    # 仕様: 1行解析時は自動判断をユーザ判断にもセットする（合計勤務時間設定後に計算）
    aj = _auto_judgment_symbol(row)
    row["auto_judgment"] = aj
    row["user_judgment_company"] = aj
    # 読取列は解析から抜いた文字列をそのまま用いる（60進/10進の別は _work_hours_string_to_decimal 側のルール）
    row["labor_read_display"] = (th_raw or "").strip() or "（なし）"
    row["seal_in_doc"] = _seal_norm or _extract_seal_in_from_document(
        text_for_extraction
    )
    if _filename_has_transport_expense_marker(fn):
        row.setdefault("transport_expense_raw", "（不明）")
        te_src = analysis if prefer_kintai_section else text_for_extraction
        te_raw = (ktab.get("transport") or "").strip() or _extract_transport_expense_from_document(
            te_src
        )
        if te_raw:
            row["transport_expense_raw"] = te_raw


def _row_is_excel(row: dict[str, str]) -> bool:
    ext = Path(row.get("file_name") or "").suffix.lower()
    return ext in EXCEL_SUFFIXES or (row.get("source_kind") or "").strip() == "excel"


def _target_sheet_ok_symbol(row: dict[str, str]) -> str:
    if "target_sheet_exists" not in row:
        return ""
    v = (row.get("target_sheet_exists") or "").strip()
    return v if v in ("〇", "✖") else ""


def _escape_md_table_cell(text: str) -> str:
    """Markdown表セル内用に | と改行をエスケープする。"""
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("|", "\\|")
    return text.replace("\n", "<br>")


# アップロードが 422 等で失敗した際、ワーカーのチャットを作り直す判断に使う
_UPLOAD_RECREATE_CHAT_STATUSES = frozenset(
    {400, 404, 407, 408, 409, 410, 422, 423, 424, 425, 426, 500}
)


def _extract_http_status_from_adk_exc(exc: BaseException) -> int | None:
    """NewtonX ADK が付ける FileUploadError 等から HTTP ステータスを推定。"""
    sc = getattr(exc, "status_code", None)
    if isinstance(sc, int):
        return sc
    msg = str(exc)
    m = re.search(r"に失敗:\s*(\d{3})\s*-", msg)
    if m:
        return int(m.group(1))
    return None


def _upload_http_error_should_recreate_chat(exc: BaseException) -> bool:
    code = _extract_http_status_from_adk_exc(exc)
    return code is not None and code in _UPLOAD_RECREATE_CHAT_STATUSES


SUMMARY_MD_HEADER = (
    f"| {SUMMARY_ROW_NO_COL} | {SUMMARY_UPLOAD_COL} |{SUMMARY_TARGET_SHEET_COL} | "
    f"{TARGET_FILE_NAME_COL} | "
    f"{SUMMARY_FINAL_JUDGMENT_COL} | {SUMMARY_BILLING_UPDATE_RESULT_COL} | 自動判断 | "
    f"{SUMMARY_YEAR_COL} | {SUMMARY_MONTH_COL} | "
    f"{SUMMARY_COMPANY_COL} | {SUMMARY_PERSON_COL} | {SUMMARY_EMPLOYEE_NO_COL} | "
    f"{SUMMARY_BILLING_UPDATE_HOURS_COL} | "
    "合計勤務時間（10進） | 合計勤務時間（読取） | "
    f"{SUMMARY_BILLING_UPDATE_TRANSPORT_COL} | {SUMMARY_TRANSPORT_EXPENSE_COL} | "
    "会社名比較 | 押印有無 |"
)

# 旧UI列名（JSON読み込み互換）
LEGACY_MATCH_COMPANY_COL = "会社名比較（ファイル名✖文書）"


_EMPLOYEE_NO_VALID_RE = re.compile(r"^(?:\d{7}|BP\d{5})$", re.IGNORECASE)
_DECIMAL_HOURS_VALID_RE = re.compile(r"^\d+(?:\.\d{1,2})?$")


def _is_valid_employee_no(employee_no: str) -> bool:
    t = (employee_no or "").strip()
    return bool(t and _EMPLOYEE_NO_VALID_RE.fullmatch(t))


def _is_valid_total_hours_decimal(total_hours_decimal: str) -> bool:
    """合計勤務時間（10進）が「整数 or 小数点以下2桁まで」の形式かを判定する。"""
    t = (total_hours_decimal or "").strip()
    if not t or t in ("（なし）", "不明"):
        return False
    # 表示用（8.5等）/内部用（8.50等）どちらでも許容
    return bool(_DECIMAL_HOURS_VALID_RE.fullmatch(t))


def normalize_judgment_symbol(value: str) -> str:
    """判断記号を表示用（〇/△/✖）に正規化する。"""
    t = (value or "").strip()
    if t in ("×", "✕"):
        return "✖"
    return t


def auto_judgment_symbol(row: dict[str, str]) -> str:
    """自動判断（〇/△/✖）を計算する（UI列名・内部キー両対応）。"""
    emp = (
        row.get("employee_no")
        or row.get(SUMMARY_EMPLOYEE_NO_COL)
        or row.get(LEGACY_EMPLOYEE_NO_COL)
        or ""
    ).strip()
    th = (row.get("total_hours_decimal") or row.get("合計勤務時間（10進）") or "").strip()
    mc = (
        row.get("match_company")
        or row.get("会社名比較")
        or row.get(LEGACY_MATCH_COMPANY_COL)
        or ""
    ).strip()
    return _auto_judgment_symbol(
        {
            "employee_no": emp,
            "total_hours_decimal": th,
            "match_company": mc,
            "file_name": (
                row.get("file_name")
                or row.get(TARGET_FILE_NAME_COL)
                or row.get(LEGACY_TARGET_FILE_NAME_COL)
                or row.get(LEGACY_FILE_NAME_COL)
                or ""
            ).strip(),
            "source_kind": (row.get("source_kind") or "").strip(),
            "year": (
                row.get("year")
                or row.get(SUMMARY_YEAR_COL)
                or row.get(LEGACY_YEAR_COL)
                or ""
            ).strip(),
            "month": (
                row.get("month")
                or row.get(SUMMARY_MONTH_COL)
                or row.get(LEGACY_MONTH_COL)
                or ""
            ).strip(),
            "expected_year": (row.get("expected_year") or "").strip(),
            "expected_month": (row.get("expected_month") or "").strip(),
        }
    )


def _effective_user_judgment(row: dict[str, str], auto: str | None = None) -> str:
    """表示用ユーザ判断。未設定・解析時自動判断・旧仕様（会社名比較のみ）の場合は auto を返す。"""
    aj = auto if auto is not None else auto_judgment_symbol(row)
    uj_raw = normalize_judgment_symbol(
        (
            row.get("user_judgment_company")
            or row.get(SUMMARY_FINAL_JUDGMENT_COL)
            or row.get(LEGACY_USER_JUDGMENT_COL)
            or ""
        ).strip()
    )
    if not uj_raw:
        return aj
    stored_aj = normalize_judgment_symbol((row.get("auto_judgment") or "").strip())
    if stored_aj and uj_raw == stored_aj:
        return aj
    mc = (
        row.get("match_company")
        or row.get("会社名比較")
        or row.get(LEGACY_MATCH_COMPANY_COL)
        or ""
    ).strip()
    if mc and uj_raw == mc and uj_raw != aj:
        return aj
    return uj_raw


def is_manual_user_judgment(row: dict[str, str]) -> bool:
    """最終判断が自動判断と異なる＝手動変更済み。"""
    uj_raw = normalize_judgment_symbol(
        (
            row.get("user_judgment_company")
            or row.get(SUMMARY_FINAL_JUDGMENT_COL)
            or row.get(LEGACY_USER_JUDGMENT_COL)
            or ""
        ).strip()
    )
    if not uj_raw:
        return False
    aj = auto_judgment_symbol(row)
    if uj_raw == aj:
        return False
    stored_aj = normalize_judgment_symbol((row.get("auto_judgment") or "").strip())
    if stored_aj and uj_raw == stored_aj:
        return False
    mc = (
        row.get("match_company")
        or row.get("会社名比較")
        or row.get(LEGACY_MATCH_COMPANY_COL)
        or ""
    ).strip()
    if mc and uj_raw == mc:
        return False
    return True


def _billing_file_update_result_display(row: dict[str, str]) -> str:
    return (
        row.get("billing_file_update_result")
        or row.get(SUMMARY_BILLING_UPDATE_RESULT_COL)
        or row.get(LEGACY_BILLING_UPDATE_RESULT_COL)
        or ""
    ).strip()


def _row_total_hours_decimal(row: dict[str, str]) -> str:
    return (
        row.get("total_hours_decimal")
        or row.get("合計勤務時間（10進）")
        or ""
    ).strip()


def _row_billing_update_hours_decimal(row: dict[str, str]) -> str:
    return (
        row.get("billing_update_hours_decimal")
        or row.get(SUMMARY_BILLING_UPDATE_HOURS_COL)
        or row.get(LEGACY_BILLING_UPDATE_HOURS_COL)
        or ""
    ).strip()


def _row_billing_update_transport(row: dict[str, str]) -> str:
    return (
        row.get("billing_update_transport")
        or row.get(SUMMARY_BILLING_UPDATE_TRANSPORT_COL)
        or ""
    ).strip()


def _row_total_hours_raw(row: dict[str, str]) -> str:
    return (
        row.get("total_hours_raw")
        or row.get("合計勤務時間（読取）")
        or ""
    ).strip()


def _row_employee_no(row: dict[str, str]) -> str:
    fn = (
        row.get("file_name")
        or row.get(TARGET_FILE_NAME_COL)
        or row.get(LEGACY_TARGET_FILE_NAME_COL)
        or row.get(LEGACY_FILE_NAME_COL)
        or ""
    ).strip()
    return (
        row.get("employee_no")
        or row.get(SUMMARY_EMPLOYEE_NO_COL)
        or row.get(LEGACY_EMPLOYEE_NO_COL)
        or _employee_no_from_file_name(fn)
        or ""
    ).strip()


def _clear_billing_update_columns(row: dict[str, str]) -> None:
    row["billing_update_hours_decimal"] = ""
    row["billing_update_transport"] = ""
    row[SUMMARY_BILLING_UPDATE_HOURS_COL] = ""
    row[SUMMARY_BILLING_UPDATE_TRANSPORT_COL] = ""
    row.pop(LEGACY_BILLING_UPDATE_HOURS_COL, None)


def _set_billing_update_columns(
    row: dict[str, str], *, hours: str, transport: str
) -> None:
    row["billing_update_hours_decimal"] = hours
    row["billing_update_transport"] = transport
    row[SUMMARY_BILLING_UPDATE_HOURS_COL] = hours
    row[SUMMARY_BILLING_UPDATE_TRANSPORT_COL] = transport


def _format_transport_sum(value: float) -> str:
    if math.isclose(value, round(value), abs_tol=1e-9):
        return str(int(round(value)))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def populate_billing_update_columns(rows: list[dict[str, str]]) -> int:
    """表示順の行リストについて更新用列を設定する（in-place）。

    社員番号が1件のみ: 合計勤務時間（10進）・交通費合計（読取）をコピー。
    複数件: No順先頭行に読取合算値を設定し、他行の更新用列は空にする。
    戻り値: 更新用値を設定したグループ数。
    """
    if not rows:
        return 0

    for row in rows:
        _clear_billing_update_columns(row)

    groups: dict[str, list[int]] = {}
    for idx, row in enumerate(rows):
        emp = _row_employee_no(row)
        if not _is_valid_employee_no(emp):
            continue
        groups.setdefault(emp, []).append(idx)

    updated_groups = 0
    for indices in groups.values():
        if len(indices) == 1:
            row = rows[indices[0]]
            hours = _row_total_hours_decimal(row)
            transport = _row_transport_expense_raw(row)
            if not hours and not transport:
                continue
            _set_billing_update_columns(
                row,
                hours=hours,
                transport=transport,
            )
            updated_groups += 1
            continue

        first_idx = indices[0]
        hour_sum = 0.0
        hours_ok = True
        for idx in indices:
            raw_hours = _row_total_hours_raw(rows[idx])
            parsed = _work_hours_raw_to_hours_float(raw_hours)
            if parsed is None:
                hours_ok = False
                break
            hour_sum += parsed

        transport_sum = 0.0
        transport_ok = True
        for idx in indices:
            amount, ok = _transport_amount_for_excel(_row_transport_expense_raw(rows[idx]))
            if not ok or amount is None:
                transport_ok = False
                break
            transport_sum += amount

        hours_str = _format_decimal_str(hour_sum) if hours_ok else ""
        transport_str = _format_transport_sum(transport_sum) if transport_ok else ""
        if not hours_str and not transport_str:
            continue
        _set_billing_update_columns(
            rows[first_idx],
            hours=hours_str,
            transport=transport_str,
        )
        updated_groups += 1

    return updated_groups


def _row_transport_expense_raw(row: dict[str, str]) -> str:
    return (
        row.get("transport_expense_raw")
        or row.get(SUMMARY_TRANSPORT_EXPENSE_COL)
        or ""
    ).strip()


def _normalize_employee_no_cell_value(value: object) -> str:
    """Excel A列などの社員番号セルを照合用キーに正規化する。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, (int, float)):
        if value == int(value):
            n = int(value)
            if 0 < n < 10_000_000:
                s = str(n)
                if s.isdigit() and len(s) == 7:
                    return s
                if len(s) <= 7 and s.isdigit():
                    return s.zfill(7)
    t = unicodedata.normalize("NFKC", str(value).strip())
    if not t or t == "社員番号エラー":
        return ""
    if t.isdigit() and len(t) == 7:
        return t
    if re.fullmatch(r"BP\d{5}", t, re.IGNORECASE):
        return t[:2].upper() + t[2:]
    return ""


def _hours_decimal_for_excel(value: str) -> float | None:
    t = (value or "").strip()
    if not t or not _is_valid_total_hours_decimal(t):
        return None
    try:
        return float(t)
    except ValueError:
        return None


def _transport_amount_for_excel(value: str) -> tuple[float | None, bool]:
    """旅費交通費請求金額（O列）用。空欄は 0 として扱う。(金額, 解析成功)"""
    t = (value or "").strip()
    if not t or t in ("（なし）", "不明", "（不明）"):
        return 0.0, True
    nfkc = unicodedata.normalize("NFKC", t)
    nfkc = nfkc.replace("，", "").replace(",", "").replace("、", "").replace("円", "")
    m = re.search(r"(\d+(?:\.\d+)?)", nfkc)
    if not m:
        return None, False
    try:
        return float(m.group(1)), True
    except ValueError:
        return None, False


def _build_employee_no_row_index(ws) -> dict[str, list[int]]:
    """シート A列の社員番号 → 行番号リスト（1始まり）。"""
    index: dict[str, list[int]] = {}
    max_row = ws.max_row or 0
    for row_idx in range(1, max_row + 1):
        key = _normalize_employee_no_cell_value(ws.cell(row=row_idx, column=BILLING_TS_COL_EMPLOYEE_NO).value)
        if not key:
            continue
        index.setdefault(key, []).append(row_idx)
    return index


def update_billing_engineer_ts_sheet(
    billing_path: Path,
    rows: list[dict[str, str]],
) -> list[str]:
    """最終判断〇行分のデータを請求用 Excel「エンジニアTS一覧」に反映する。

    各行について「〇」（更新成功）または「✖」（失敗）を返す（rows と同順）。
    更新前にファイル名の会社・氏名と請求シート（D/E列）の一致を確認する。
    """
    if not rows:
        return []
    path = billing_path.resolve()
    try:
        sheet_rows = _load_billing_engineer_ts_rows(path)
    except Exception:
        return ["✖"] * len(rows)
    if not sheet_rows:
        return ["✖"] * len(rows)
    suffix = path.suffix.lower()
    keep_vba = suffix in (".xlsm", ".xltm")
    wb = load_workbook(path, keep_vba=keep_vba)
    try:
        if BILLING_ENGINEER_TS_SHEET_NAME not in wb.sheetnames:
            return ["✖"] * len(rows)
        ws = wb[BILLING_ENGINEER_TS_SHEET_NAME]
        emp_rows = _build_employee_no_row_index(ws)
        results: list[str] = []
        for row in rows:
            file_name = (row.get("file_name") or "").strip()
            if not file_name:
                results.append("✖")
                continue
            billing_emp = lookup_employee_no_in_billing_file(
                path, file_name, sheet_rows=sheet_rows
            )
            if not billing_emp:
                results.append(BILLING_UPDATE_NO_TARGET_RECORD)
                continue
            emp = _normalize_employee_no_cell_value(billing_emp)
            if not _is_valid_employee_no(emp):
                results.append(BILLING_UPDATE_NO_TARGET_RECORD)
                continue
            targets = emp_rows.get(emp)
            if not targets:
                results.append(BILLING_UPDATE_NO_TARGET_RECORD)
                continue
            hours = _hours_decimal_for_excel(_row_billing_update_hours_decimal(row))
            if hours is None:
                results.append("✖")
                continue
            transport, transport_ok = _transport_amount_for_excel(
                _row_billing_update_transport(row)
            )
            if not transport_ok or transport is None:
                results.append("✖")
                continue
            try:
                for row_idx in targets:
                    ws.cell(row=row_idx, column=BILLING_TS_COL_NORMAL_HOURS).value = hours
                    ws.cell(row=row_idx, column=BILLING_TS_COL_TRANSPORT_AMOUNT).value = transport
                results.append("〇")
            except Exception:
                results.append("✖")
        wb.save(path)
        return results
    finally:
        wb.close()


def _auto_judgment_symbol(row: dict[str, str]) -> str:
    """自動判断（〇/△/✖）を計算する。

    仕様:
      - 読み取った年・月とコンボ（expected_year/month）が不一致 → ✖
      - 〇: 年・月一致、社員番号が有効、合計勤務時間(10進)が有効、会社名比較が〇
      - △: 上記と同様だが会社名比較が△（画像/PDF・Excel 共通）
      - ✖: 年・月不一致、または社員番号・勤務時間・会社名比較のいずれかが不適合
    """
    if not _year_month_matches_expected(row):
        return "✖"
    emp = (row.get("employee_no") or "").strip()
    th = (row.get("total_hours_decimal") or "").strip()
    mc = (row.get("match_company") or "").strip()
    if not (_is_valid_employee_no(emp) and _is_valid_total_hours_decimal(th)):
        return "✖"
    if mc == "〇":
        return "〇"
    if mc == "△":
        return "△"
    return "✖"


def _upload_image_with_retries(
    wc: NewtonXClient,
    *,
    chat_uid_holder: dict[str, str],
    upload_src: str | Path,
    file_name: str,
    recreate_chat_fn: Callable[[BaseException], bool] | None = None,
    log_emit: Callable[[str], None] | None = None,
) -> tuple[str | None, bool]:
    """upload_image を試行する。422 などの場合は recreate_chat_fn でチャット刷新後・同一試行で再アップロード。"""
    last: str | None = None
    lg = log_emit if log_emit is not None else print

    def _recover_if_applicable(exc: BaseException, *, recovered_already: bool) -> bool:
        if not recreate_chat_fn or recovered_already:
            return False
        if not _upload_http_error_should_recreate_chat(exc):
            return False
        code = _extract_http_status_from_adk_exc(exc)
        code_s = str(code) if code is not None else "?"
        lg(f"画像アップロード失敗 HTTP {code_s} のためチャットを再作成して再試行します。")
        return bool(recreate_chat_fn(exc))

    for attempt in range(UPLOAD_MAX_RETRIES + 1):
        if attempt > 0 and UPLOAD_RETRY_DELAY_SEC > 0:
            time.sleep(UPLOAD_RETRY_DELAY_SEC)
        recovered_round = False
        while True:
            try:
                last = wc.upload_image(
                    chat_uid=chat_uid_holder["chat_uid"],
                    file_path=upload_src,
                    file_name=file_name,
                )
                if last:
                    return last, True
                break
            except FileUploadError as e:
                if _extract_http_status_from_adk_exc(e) is None:
                    lg(f"画像アップロード: リトライ不能な検証／IOエラーです — {e}")
                    return None, False
                if _recover_if_applicable(e, recovered_already=recovered_round):
                    recovered_round = True
                    continue
                last = None
                break
    return last, False


def _upload_document_with_retries(
    wc: NewtonXClient,
    *,
    chat_uid_holder: dict[str, str],
    file_path: str,
    file_name: str,
    recreate_chat_fn: Callable[[BaseException], bool] | None = None,
    log_emit: Callable[[str], None] | None = None,
) -> tuple[bool, bool]:
    """upload_document を試行する。チャット刷新ロジックは画像と同一。"""
    last_ok = False
    lg = log_emit if log_emit is not None else print

    def _recover_if_applicable(exc: BaseException, *, recovered_already: bool) -> bool:
        if not recreate_chat_fn or recovered_already:
            return False
        if not _upload_http_error_should_recreate_chat(exc):
            return False
        code = _extract_http_status_from_adk_exc(exc)
        code_s = str(code) if code is not None else "?"
        lg(f"PDFアップロード失敗 HTTP {code_s} のためチャットを再作成して再試行します。")
        return bool(recreate_chat_fn(exc))

    for attempt in range(UPLOAD_MAX_RETRIES + 1):
        if attempt > 0 and UPLOAD_RETRY_DELAY_SEC > 0:
            time.sleep(UPLOAD_RETRY_DELAY_SEC)
        recovered_round = False
        while True:
            try:
                last_ok = wc.upload_document(
                    chat_uid=chat_uid_holder["chat_uid"],
                    file_path=file_path,
                    file_name=file_name,
                )
                if last_ok:
                    return True, True
                break
            except FileUploadError as e:
                if _extract_http_status_from_adk_exc(e) is None:
                    lg(f"PDFアップロード: リトライ不能な検証／IOエラーです — {e}")
                    return False, False
                if _recover_if_applicable(e, recovered_already=recovered_round):
                    recovered_round = True
                    continue
                last_ok = False
                break
    return last_ok, False


def _upload_ok_symbol(row: dict[str, str]) -> str:
    """JSON に upload_ok が無い既存データは空欄。無効値も空欄に寄せる。"""
    if "upload_ok" not in row:
        return ""
    v = (row.get("upload_ok") or "").strip()
    return v if v in ("〇", "✖") else ""


def _year_month_from_data_dir(data_dir: Path) -> tuple[str, str]:
    """data_dir（.../YYYY年M月/...）から年・月文字列を抽出する。"""
    for part in reversed(data_dir.resolve().parts):
        m = re.fullmatch(r"(\d{4})年(\d{1,2})月", part)
        if m:
            return m.group(1), str(int(m.group(2)))
    return "", ""


def _one_summary_data_line(r: dict[str, str], *, row_no: int | None = None) -> str:
    """19列1行分（集計用）。Excel 行は対象シート有無のみを埋め、AI読取列は空欄にする。"""
    is_excel = _row_is_excel(r)
    u_sym = _escape_md_table_cell(_upload_ok_symbol(r))
    ts = _escape_md_table_cell(_target_sheet_ok_symbol(r))
    fn = _escape_md_table_cell(r.get("file_name", ""))

    aj = auto_judgment_symbol(r)
    r["auto_judgment"] = aj
    uj = _effective_user_judgment(r, aj)
    r["user_judgment_company"] = uj
    uj = _escape_md_table_cell(uj)
    bur = _escape_md_table_cell(_billing_file_update_result_display(r))
    aj = _escape_md_table_cell(aj)
    yy = _escape_md_table_cell((r.get("year") or "").strip())
    mm = _escape_md_table_cell((r.get("month") or "").strip())

    co1 = _escape_md_table_cell(
        ((r.get("name_company_1") or "").strip() or ("" if is_excel else "不明"))
    )
    pe = _escape_md_table_cell(
        ((r.get("name_person_from_doc") or "").strip() or ("" if is_excel else "不明"))
    )
    emp = _escape_md_table_cell(
        (r.get("employee_no") or "").strip() or ""
    )
    buh = _escape_md_table_cell(
        _decimal_for_table_display(_row_billing_update_hours_decimal(r))
        if _row_billing_update_hours_decimal(r)
        else ""
    )
    th = _escape_md_table_cell(
        ((
            _decimal_for_table_display((r.get("total_hours_decimal") or "").strip())
        ) if (r.get("total_hours_decimal") or "").strip() else ("" if is_excel else "（なし）"))
    )
    lr = _escape_md_table_cell(
        ((r.get("total_hours_raw") or "").strip() or ("" if is_excel else "（なし）"))
    )
    but = _escape_md_table_cell(_row_billing_update_transport(r))
    te = _escape_md_table_cell((r.get("transport_expense_raw") or "").strip())
    mc = _escape_md_table_cell((r.get("match_company") or ("" if is_excel else "✖")))
    se = _escape_md_table_cell(
        "" if is_excel else ((r.get("seal_in_doc") or "").strip() or "不明")
    )
    no = _escape_md_table_cell(str(row_no) if row_no is not None else "")
    return (
        f"| {no} | {u_sym} | {ts} | {fn} | {uj} | {bur} | {aj} | {yy} | {mm} | "
        f"{co1} | {pe} | {emp} | {buh} | {th} | {lr} | {but} | {te} | {mc} | {se} |"
    )


def _company_match_counts(results: list[dict[str, str]]) -> tuple[int, int]:
    """会社名比較の集計用に (〇扱い件数, 実行済件数) を返す。△は〇側に含める。"""
    ok_count = 0
    processed_count = 0
    for row in results:
        processed_count += 1
        symbol = (row.get("match_company") or "").strip()
        if symbol in ("〇", "△"):
            ok_count += 1
    return ok_count, processed_count


def _company_match_ratio_lines(
    results: list[dict[str, str]],
    total_target_count: int | None = None,
) -> list[str]:
    """会社名比較の〇率を、画像・PDF・Excel の全処理件数ベースで返す。"""
    ok_count, processed_count = _company_match_counts(results)
    target_count = total_target_count if total_target_count is not None else processed_count
    if processed_count == 0:
        return ["会社名比較 〇率: 対象データなし（実行済 0 / 対象 0）", ""]

    ratio = (ok_count / processed_count) * 100
    return [
        f"会社名比較 〇率: {ratio:.1f}% （〇扱い {ok_count}件 / 実行済 {processed_count}件 / 対象 {target_count}件）",
        "",
    ]


def _company_match_ratio_progress_line(
    ok_count: int,
    processed_count: int,
    total_target_count: int,
) -> str:
    """逐次ログ用の会社名比較〇率文字列を返す。"""
    if processed_count <= 0:
        return (
            "会社名比較 〇率(途中経過): "
            f"対象データなし（実行済 0 / 対象 {total_target_count}）"
        )
    ratio = (ok_count / processed_count) * 100
    return (
        f"会社名比較 〇率(途中経過): {ratio:.1f}% "
        f"（〇扱い {ok_count}件 / 実行済 {processed_count}件 / 対象 {total_target_count}件）"
    )


def _summary_table_md_lines(
    results: list[dict[str, str]],
    total_target_count: int | None = None,
) -> list[str]:
    """
    解析結果.md 用: ヘッダ1行＋区切り行なし。データは各1行。
    """
    return _company_match_ratio_lines(results, total_target_count=total_target_count) + [SUMMARY_MD_HEADER] + [
        _one_summary_data_line(r, row_no=i + 1) for i, r in enumerate(results)
    ]


def summary_header_cells() -> tuple[str, ...]:
    """SUMMARY_MD_HEADER から見出し要素を返す（Treeview 列名用）。"""
    raw = SUMMARY_MD_HEADER.strip()
    if raw.startswith("|"):
        raw = raw[1:]
    if raw.endswith("|"):
        raw = raw[:-1]
    return tuple(c.strip() for c in raw.split("|"))


def row_display_values(
    r: dict[str, str],
    *,
    sync_user_judgment_to_auto: bool = False,
    row_no: int | None = None,
) -> tuple[str, ...]:
    """グリッド表示用（Markdown エスケープなし）。 _one_summary_data_line と同一ルール。"""
    is_excel = _row_is_excel(r)
    up_sym = _upload_ok_symbol(r)
    ts = _target_sheet_ok_symbol(r)

    aj = auto_judgment_symbol(r)
    r["auto_judgment"] = aj
    if sync_user_judgment_to_auto:
        uj = aj
    else:
        uj = _effective_user_judgment(r, aj)
    r["user_judgment_company"] = uj
    bur = _billing_file_update_result_display(r)

    fn = r.get("file_name", "") or ""
    yy = (r.get("year") or "").strip()
    mm = (r.get("month") or "").strip()
    co1 = ((r.get("name_company_1") or "").strip() or ("" if is_excel else "不明"))
    pe = ((r.get("name_person_from_doc") or "").strip() or ("" if is_excel else "不明"))
    emp = (r.get("employee_no") or "").strip() or ""
    buh_raw = _row_billing_update_hours_decimal(r)
    buh = _decimal_for_table_display(buh_raw) if buh_raw else ""
    th = (
        _decimal_for_table_display((r.get("total_hours_decimal") or "").strip())
        if (r.get("total_hours_decimal") or "").strip()
        else ("" if is_excel else "（なし）")
    )
    lr = ((r.get("total_hours_raw") or "").strip() or ("" if is_excel else "（なし）"))
    but = _row_billing_update_transport(r)
    te = (r.get("transport_expense_raw") or "").strip()
    mc = (r.get("match_company") or ("" if is_excel else "✖"))
    se = "" if is_excel else ((r.get("seal_in_doc") or "").strip() or "不明")
    no = str(row_no) if row_no is not None else ""
    return (
        no,
        up_sym,
        ts,
        fn,
        uj,
        bur,
        aj,
        yy,
        mm,
        co1,
        pe,
        emp,
        buh,
        th,
        lr,
        but,
        te,
        mc,
        se,
    )


def run_analysis(
    client: NewtonXClient,
    assistant_uid: str,
    data_dir: Path,
    *,
    save_md_path: Path | None = None,
    on_log: Callable[[str], None] | None = None,
    emit_progress_md_rows: bool = True,
    on_file_started: Callable[[str], None] | None = None,
    on_file_progress: Callable[[int, int], None] | None = None,
    on_row_completed: Callable[[dict[str, str]], None] | None = None,
    cancel_event=None,
    skip_file_names: set[str] | None = None,
    target_file_names: set[str] | None = None,
    parallel_chats: int = DEFAULT_PARALLEL_ANALYSIS_CHATS,
    expected_year: int | str | None = None,
    expected_month: int | str | None = None,
) -> list[dict[str, str]]:
    """
    指定フォルダ直下の画像・PDF・Excel を名前順で処理し、結果行のリストを返す。
    NewtonX では parallel_chats 本のワーカーでファイルを並列処理するが、
    チャットはワーカー固定ではなく **1ファイルごとに新規作成** し、
    処理完了後に削除する。
    各ワーカーは独自の NewtonXClient で API を呼び出す。
    アップロードは、失敗のたびに最大 UPLOAD_MAX_RETRIES 回まで再試行します（総試行は多くとも 1+UPLOAD_MAX_RETRIES 回）。
    Excel は生成AIへアップロードせず、対象シート名の有無のみを判定する。
    save_md_path が None のときは cwd に 解析結果.md を出力する。
    emit_progress_md_rows が False のとき、コンソール向け Markdown 行は on_log に流さない。
    on_file_started: ワーカーが各ファイルの処理に着手したタイミングでファイル名を通知する。
        GUI などで「現在処理中の行」をハイライトする用途を想定。
    on_file_progress: 処理が終わったファイル数を (実行済み数, 対象総数) で通知する（各ファイルの試行の末尾で1回）。
    on_row_completed: 解析結果またはアップロード失敗行を1件 results に載せた直後に呼ぶ。
    cancel_event: threading.Event 互換。set() されたら以降の解析を中断する（処理中の1ファイルは止まらず、次の境界で止まる）。
    skip_file_names: ここに含まれるファイル名は解析処理自体をスキップする（progress は進める）。
    target_file_names: ここに含まれるファイル名だけを解析対象にする。None の場合は全対象ファイル。
    """
    log = on_log if on_log is not None else print
    row_year, row_month = _year_month_from_data_dir(data_dir)
    if expected_year is not None:
        exp_year_str = str(int(expected_year))
    else:
        exp_year_str = row_year
    if expected_month is not None:
        exp_month_str = str(int(expected_month))
    elif row_month:
        exp_month_str = row_month
    else:
        exp_month_str = ""

    if not data_dir.is_dir():
        raise FileNotFoundError(f"フォルダが存在しません: {data_dir}")

    skip_names = {str(x).strip() for x in (skip_file_names or set()) if str(x).strip()}
    target_names = {str(x).strip() for x in (target_file_names or set()) if str(x).strip()}

    def _is_target_name(p: Path) -> bool:
        return (not target_names) or p.name in target_names

    image_files = sorted(
        p for p in data_dir.iterdir()
        if p.is_file() and p.suffix.lower() in IMAGE_SUFFIXES and p.name not in skip_names and _is_target_name(p)
    )
    pdf_files = sorted(
        p for p in data_dir.iterdir()
        if p.is_file() and p.suffix.lower() == PDF_SUFFIX and p.name not in skip_names and _is_target_name(p)
    )
    excel_files = sorted(
        p for p in data_dir.iterdir()
        if p.is_file() and p.suffix.lower() in EXCEL_SUFFIXES and p.name not in skip_names and _is_target_name(p)
    )
    if not image_files and not pdf_files and not excel_files:
        if target_names:
            raise ValueError(f"指定された対象ファイルが見つかりません: {', '.join(sorted(target_names))}")
        raise ValueError(f"画像・PDF・Excelファイルが見つかりません: {data_dir}")

    total_files = len(image_files) + len(pdf_files) + len(excel_files)
    progress_done = 0
    progress_lock = threading.Lock()

    def is_cancelled() -> bool:
        try:
            return bool(cancel_event is not None and cancel_event.is_set())
        except Exception:
            return False

    def notify_file_finished() -> None:
        nonlocal progress_done
        with progress_lock:
            progress_done += 1
            cur = progress_done
        if on_file_progress is not None:
            on_file_progress(cur, total_files)

    if on_file_progress is not None:
        on_file_progress(0, total_files)

    if is_cancelled():
        log("中断: 解析を開始する前にキャンセルされました")
        return []

    target_folder_name = "業務課集計"
    folders = client.get_folders()
    matched = next(
        (
            f
            for f in folders
            if (f.get("name") or "").strip() == target_folder_name
        ),
        None,
    )
    if matched is not None:
        raw_id = (
            matched.get("uid")
            if matched.get("uid") is not None
            else matched.get("id")
        )
        folder_uid = str(raw_id) if raw_id is not None else None
        if not folder_uid:
            raise RuntimeError(
                "NewtonX のフォルダ一覧に該当がありますが、フォルダ ID を取得できませんでした。"
            )
        log(f"既存フォルダを使用します: {target_folder_name} ({folder_uid})")
    else:
        folder_uid = client.create_folder(target_folder_name)
        if not folder_uid:
            raise RuntimeError("NewtonX 上でフォルダの作成に失敗しました。")
        log(f"フォルダが作成されました: {folder_uid}")

    n_workers = int(parallel_chats)
    if n_workers < 1:
        n_workers = DEFAULT_PARALLEL_ANALYSIS_CHATS
    elif n_workers > PARALLEL_WORKERS_MAX:
        n_workers = PARALLEL_WORKERS_MAX

    summary_header_done = False
    summary_lock = threading.Lock()
    company_ratio_lock = threading.Lock()
    company_match_ok_count = 0
    company_match_processed_count = 0

    def emit_summary_row_md(row_dict: dict[str, str]) -> None:
        nonlocal summary_header_done
        if not emit_progress_md_rows:
            return
        with summary_lock:
            if not summary_header_done:
                log(SUMMARY_MD_HEADER)
                summary_header_done = True
            log(_one_summary_data_line(row_dict))

    def emit_company_match_ratio_progress(row_dict: dict[str, str]) -> None:
        nonlocal company_match_ok_count, company_match_processed_count
        symbol = (row_dict.get("match_company") or "").strip()
        with company_ratio_lock:
            company_match_processed_count += 1
            if symbol in ("〇", "△"):
                company_match_ok_count += 1
            line = _company_match_ratio_progress_line(
                company_match_ok_count,
                company_match_processed_count,
                total_files,
            )
        log(line)

    tasks_ordered: list[tuple[str, Path]] = [
        ("image", p) for p in image_files
    ] + [("pdf", p) for p in pdf_files] + [("excel", p) for p in excel_files]

    buckets: list[list[tuple[str, Path]]] = [[] for _ in range(n_workers)]
    for i, task in enumerate(tasks_ordered):
        buckets[i % n_workers].append(task)

    bucket_results: list[list[dict[str, str]]] = [[] for _ in range(n_workers)]
    worker_exc: list[BaseException | None] = [None] * n_workers

    def attach_year_month(row: dict[str, str]) -> None:
        if exp_year_str:
            row["expected_year"] = exp_year_str
        if exp_month_str:
            row["expected_month"] = exp_month_str
        # Excel の年・月は A7/D7（_extract_excel_target_sheet_row）から設定する

    def append_upload_failure_row(
        worker_ix: int,
        file_path: Path,
        *,
        prefer_pdf_section: bool,
        message: str,
    ) -> None:
        row = {
            "upload_ok": "✖",
            "file_name": file_path.name,
            "resolved_path": str(file_path.resolve()),
            "analysis": message,
        }
        attach_year_month(row)
        _enrich_with_match_scores(
            row,
            row["analysis"],
            prefer_kintai_section=prefer_pdf_section,
        )
        bucket_results[worker_ix].append(row)
        emit_summary_row_md(row)
        emit_company_match_ratio_progress(row)
        if on_row_completed is not None:
            on_row_completed(row)

    def worker_loop(worker_idx: int, tasks: list[tuple[str, Path]]) -> None:
        wc = create_client()

        def create_file_chat(title_suffix: str) -> str | None:
            try:
                title = f"業務課集計 {title_suffix}"
                create_chat_in_folder = getattr(wc, "create_chat_in_folder", None)
                if callable(create_chat_in_folder):
                    try:
                        nc = create_chat_in_folder(
                            assistant_uid=assistant_uid,
                            folder_uid=folder_uid,
                            title=title,
                        )
                    except TypeError:
                        nc = create_chat_in_folder(
                            assistant_uid,
                            folder_uid,
                            title,
                        )
                else:
                    nc = wc.create_chat(
                        assistant_uid=assistant_uid,
                        title=title,
                    )
                if not nc:
                    return None
                try:
                    wc.move_chat_to_folder(chat_uid=nc, folder_uid=folder_uid)
                except Exception as me:
                    log(
                        f"ワーカー {worker_idx + 1}: "
                        f"チャットのフォルダ移動に警告 — {me}"
                    )
                return str(nc).strip()
            except Exception as e:
                log(f"ワーカー {worker_idx + 1}: チャット作成処理で異常 — {e}")
                return None

        def delete_file_chat(chat_uid: str | None) -> None:
            """解析用に作成したチャットを後片付けする。

            UI のステータス欄へ on_log が流れるため、削除失敗はユーザーにとってノイズになりやすい。
            ここでは削除失敗/例外は握りつぶし（静かに無視）とし、成功時のみログを残す。
            """
            if not DELETE_CHAT_AFTER_ANALYSIS:
                return
            if not chat_uid:
                return
            try:
                success = wc.delete_chat(chat_uid)
                if success:
                    log(f"チャットが削除されました: {chat_uid}")
            except Exception:
                # 削除失敗は後続処理に影響しないため、表示・停止させない
                return

        def signal_fatal(exc: BaseException) -> None:
            worker_exc[worker_idx] = exc
            try:
                if cancel_event is not None:
                    cancel_event.set()
            except Exception:
                pass

        for kind, file_path in tasks:
            if is_cancelled():
                log(
                    f"中断: ワーカー {worker_idx + 1} がキャンセルにより停止しました"
                )
                break

            # GUI 向け: 「今からこのファイルを処理する」通知（並列ワーカーの場合は同時に複数呼ばれ得る）
            if on_file_started is not None:
                try:
                    on_file_started(file_path.name)
                except Exception:
                    # 通知失敗は解析処理自体に影響させない
                    pass

            chat_holder: dict[str, str] = {"chat_uid": ""}
            try:
                if kind == "excel":
                    row_excel = _extract_excel_target_sheet_row(file_path)
                    attach_year_month(row_excel)
                    bucket_results[worker_idx].append(row_excel)
                    emit_summary_row_md(row_excel)
                    emit_company_match_ratio_progress(row_excel)
                    if on_row_completed is not None:
                        on_row_completed(row_excel)
                    if row_excel.get("analysis"):
                        log(f"Excel処理警告: {file_path.name}: {row_excel['analysis']}")
                elif kind == "image":
                    try:
                        upload_src, tmp_upload = _resolve_upload_path(file_path)
                    except (
                        OSError,
                        RuntimeError,
                        ValueError,
                        UnidentifiedImageError,
                    ) as e:
                        log(
                            f"スキップ: 画像の読み込み／圧縮に失敗しました — "
                            f"{file_path.name}: {e}"
                        )
                        append_upload_failure_row(
                            worker_idx,
                            file_path,
                            prefer_pdf_section=False,
                            message="（画像の読み込み／圧縮に失敗しました）",
                        )
                    else:
                        if is_cancelled():
                            log("中断: 画像のアップロード前にキャンセルされました")
                            break
                        tmp_upload_path: Path | None = tmp_upload
                        try:
                            chat_uid = create_file_chat(f"#{worker_idx + 1} {file_path.name}")
                            if not chat_uid:
                                raise RuntimeError(f"NewtonX 上でチャットの作成に失敗しました — {file_path.name}")
                            chat_holder["chat_uid"] = chat_uid
                            image_id, upload_succeeded = _upload_image_with_retries(
                                wc,
                                chat_uid_holder=chat_holder,
                                upload_src=upload_src,
                                file_name=file_path.name,
                                recreate_chat_fn=None,
                                log_emit=log,
                            )
                            if not image_id:
                                log(
                                    f"スキップ: アップロードに失敗しました（最大 "
                                    f"{UPLOAD_MAX_RETRIES} 回までリトライ）— "
                                    f"{file_path.name}"
                                )
                                append_upload_failure_row(
                                    worker_idx,
                                    file_path,
                                    prefer_pdf_section=False,
                                    message=(
                                        f"（画像アップロード失敗／{UPLOAD_MAX_RETRIES}回リトライまで）"
                                    ),
                                )
                            else:
                                if is_cancelled():
                                    log(
                                        "中断: 画像の解析要求前にキャンセルされました"
                                    )
                                    break
                                def analyze_image_once() -> tuple[dict[str, str] | None, str]:
                                    raw = wc.send_message(
                                        chat_uid=chat_holder["chat_uid"],
                                        message=_build_check_message(file_path.name),
                                        image_ids=[image_id],
                                    )
                                    text = raw or ""
                                    return _parse_analysis_json_response(text), text

                                if is_cancelled():
                                    log(
                                        "中断: 画像の解析応答待ち後にキャンセルされました"
                                    )
                                    break
                                row = {
                                    "upload_ok": "〇" if upload_succeeded else "✖",
                                    "file_name": file_path.name,
                                    "resolved_path": str(file_path.resolve()),
                                    "analysis": "",
                                }
                                attach_year_month(row)
                                analysis_json, raw_response = analyze_image_once()
                                _apply_analysis_response_to_row(
                                    row, analysis_json, raw_response
                                )
                                bucket_results[worker_idx].append(row)
                                emit_summary_row_md(row)
                                emit_company_match_ratio_progress(row)
                                if on_row_completed is not None:
                                    on_row_completed(row)
                        finally:
                            delete_file_chat(chat_holder.get("chat_uid") or None)
                            if tmp_upload_path is not None:
                                tmp_upload_path.unlink(missing_ok=True)
                else:
                    try:
                        upload_src, tmp_upload_paths = _pdf_to_png_for_upload(
                            file_path
                        )
                    except (
                        OSError,
                        RuntimeError,
                        ValueError,
                    ) as e:
                        log(
                            f"スキップ: PDFの画像変換に失敗しました — "
                            f"{file_path.name}: {e}"
                        )
                        append_upload_failure_row(
                            worker_idx,
                            file_path,
                            prefer_pdf_section=True,
                            message="（PDFの画像変換に失敗しました）",
                        )
                    else:
                        if is_cancelled():
                            log("中断: PDFの画像アップロード前にキャンセルされました")
                            break
                        try:
                            chat_uid = create_file_chat(
                                f"#{worker_idx + 1} {file_path.name}"
                            )
                            if not chat_uid:
                                raise RuntimeError(
                                    f"NewtonX 上でチャットの作成に失敗しました — "
                                    f"{file_path.name}"
                                )
                            chat_holder["chat_uid"] = chat_uid
                            upload_file_name = Path(upload_src).name
                            png_kb = Path(upload_src).stat().st_size / 1024
                            log(
                                f"PDF→PNG変換完了: {file_path.name} → "
                                f"{upload_file_name} ({png_kb:.0f} KB)"
                            )
                            image_id, upload_succeeded = _upload_image_with_retries(
                                wc,
                                chat_uid_holder=chat_holder,
                                upload_src=upload_src,
                                file_name=upload_file_name,
                                recreate_chat_fn=None,
                                log_emit=log,
                            )
                            if not image_id:
                                log(
                                    f"スキップ: PDF（画像変換）のアップロードに失敗しました（最大 "
                                    f"{UPLOAD_MAX_RETRIES} 回までリトライ）— "
                                    f"{file_path.name}"
                                )
                                append_upload_failure_row(
                                    worker_idx,
                                    file_path,
                                    prefer_pdf_section=True,
                                    message=(
                                        f"（PDF画像アップロード失敗／"
                                        f"{UPLOAD_MAX_RETRIES}回リトライまで）"
                                    ),
                                )
                            else:
                                if is_cancelled():
                                    log(
                                        "中断: PDF（画像変換）の解析要求前に"
                                        "キャンセルされました"
                                    )
                                    break

                                def analyze_pdf_as_image_once() -> (
                                    tuple[dict[str, str] | None, str]
                                ):
                                    raw = wc.send_message(
                                        chat_uid=chat_holder["chat_uid"],
                                        message=_build_check_message(file_path.name),
                                        image_ids=[image_id],
                                    )
                                    text = raw or ""
                                    return _parse_analysis_json_response(text), text

                                if is_cancelled():
                                    log(
                                        "中断: PDF（画像変換）の解析応答待ち後に"
                                        "キャンセルされました"
                                    )
                                    break
                                row2 = {
                                    "upload_ok": "〇" if upload_succeeded else "✖",
                                    "file_name": file_path.name,
                                    "resolved_path": str(file_path.resolve()),
                                    "analysis": "",
                                }
                                attach_year_month(row2)
                                analysis_json, raw_response = analyze_pdf_as_image_once()
                                _apply_analysis_response_to_row(
                                    row2,
                                    analysis_json,
                                    raw_response,
                                )
                                bucket_results[worker_idx].append(row2)
                                emit_summary_row_md(row2)
                                emit_company_match_ratio_progress(row2)
                                if on_row_completed is not None:
                                    on_row_completed(row2)
                        finally:
                            delete_file_chat(chat_holder.get("chat_uid") or None)
                            for tmp_path in tmp_upload_paths:
                                tmp_path.unlink(missing_ok=True)
            except BaseException as e:
                signal_fatal(e)
                break
            finally:
                notify_file_finished()

    threads = [
        threading.Thread(
            target=worker_loop,
            args=(wi, buckets[wi]),
            name=f"kintai-worker-{wi}",
            daemon=False,
        )
        for wi in range(n_workers)
    ]
    for t in threads:
        t.start()
    for t in threads:
        t.join()

    results = []
    for wi in range(n_workers):
        results.extend(bucket_results[wi])

    def _row_kind_order(row: dict[str, str]) -> tuple[int, str]:
        fn = row.get("file_name") or ""
        ext = Path(fn).suffix.lower()
        if ext in IMAGE_SUFFIXES:
            kind_order = 0
        elif ext == PDF_SUFFIX:
            kind_order = 1
        elif ext in EXCEL_SUFFIXES:
            kind_order = 2
        else:
            kind_order = 9
        return (kind_order, fn)

    results.sort(key=_row_kind_order)

    fatal = next((e for e in worker_exc if e is not None), None)
    if fatal is not None:
        raise fatal

    if is_cancelled():
        log("中断: 結果ファイルの出力前にキャンセルされました")
        return results

    output_md = save_md_path if save_md_path is not None else Path.cwd() / "解析結果.md"
    if results:
        lines_content = "\n".join(
            _summary_table_md_lines(results, total_target_count=total_files)
        )
    else:
        lines_content = "（アップロード・解析に成功した画像・PDFがありませんでした）"
    output_md.write_text(lines_content + "\n", encoding="utf-8")
    log(f"\n解析結果を保存しました: {output_md.resolve()}")
    return results
